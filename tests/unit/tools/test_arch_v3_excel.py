"""Schema-v3 Excel adapter control tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from otdev.tools._arch.v3.excel import (
    WorkbookError,
    generate_template,
    import_workbook,
    read_workbook,
    write_workbook,
)
from otdev.tools._arch.v3.model import (
    Architecture,
    Code,
    Component,
    Container,
    Subsystem,
    System,
)
from otdev.tools._arch.v3.payload import build_payload
from otdev.tools._arch.v3.validate import validate
from otdev.tools._arch.v3.yamlio import dump_architecture, load_architecture

pytestmark = [pytest.mark.unit, pytest.mark.tools]


def test_acme_model_round_trip(tmp_path: Path) -> None:
    fixture = Path("tests/unit/tools/fixtures/arch/acme.yaml")
    architecture = load_architecture(fixture)
    workbook = tmp_path / "acme.xlsx"

    write_workbook(architecture, workbook)

    assert read_workbook(workbook) == architecture


def test_failed_import_leaves_yaml_byte_identical(tmp_path: Path) -> None:
    workbook_path = tmp_path / "broken.xlsx"
    yaml_path = tmp_path / "architecture.yaml"
    original = b"existing canonical bytes\n"
    yaml_path.write_bytes(original)
    generate_template(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Containers"]
    sheet["A2"] = "orphan"
    sheet["B2"] = "Orphan"
    sheet["C2"] = "missing-system"
    workbook.save(workbook_path)
    workbook.close()

    with pytest.raises(WorkbookError, match="unresolved_parent"):
        import_workbook(workbook_path, yaml_path)

    assert yaml_path.read_bytes() == original


def test_import_assigns_blank_ids_in_sheet_order(tmp_path: Path) -> None:
    workbook_path = tmp_path / "ids.xlsx"
    yaml_path = tmp_path / "architecture.yaml"
    generate_template(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Systems"]
    sheet.append([None, "First generated"])
    sheet.append(["s-0007", "Existing"])
    sheet.append([None, "Second generated"])
    workbook.save(workbook_path)
    workbook.close()

    result = import_workbook(workbook_path, yaml_path)

    assert result["assigned_ids"] == {
        "systems": [(0, "s-0008"), (2, "s-0009")]
    }
    assert [row.id for row in load_architecture(yaml_path).systems] == [
        "s-0008",
        "s-0007",
        "s-0009",
    ]


def test_code_and_container_round_trip(tmp_path: Path) -> None:
    architecture = Architecture(
        schema_version=3,
        milestones=[],
        systems=[System(id="s-0001", name="Platform")],
        subsystems=[],
        containers=[
            Container(id="c-0001", name="Runtime", parent="s-0001"),
            Container(id="c-0002", name="Worker", parent="s-0001"),
        ],
        components=[
            Component(id="cp-0001", name="Handler", container="c-0002")
        ],
        code=[Code(id="cd-0001", name="Module", component="cp-0001")],
        users=[],
        interfaces=[],
        relationships=[],
    )
    workbook = tmp_path / "nested.xlsx"

    write_workbook(architecture, workbook)

    assert read_workbook(workbook) == architecture


def test_subsystem_yaml_and_excel_round_trip(tmp_path: Path) -> None:
    architecture = Architecture(
        schema_version=3,
        milestones=[],
        systems=[System(id="s-0001", name="Platform")],
        subsystems=[
            Subsystem(id="ss-0001", name="Commerce", parent="s-0001")
        ],
        containers=[Container(id="c-0001", name="Storefront", parent="ss-0001")],
        components=[],
        code=[],
        users=[],
        interfaces=[],
        relationships=[],
    )
    yaml_path = tmp_path / "architecture.yaml"
    workbook_path = tmp_path / "architecture.xlsx"

    dump_architecture(architecture, yaml_path)
    write_workbook(architecture, workbook_path)

    assert load_architecture(yaml_path) == architecture
    assert read_workbook(workbook_path) == architecture


def test_theme_round_trips_yaml_excel_payload_and_reports_located_errors(
    tmp_path: Path,
) -> None:
    theme_yaml = """\
schema_version: 3
milestones:
  - id: m1
    name: Milestone
timelines:
  - id: roadmap
    milestones: [m1]
theme:
  kinds:
    system: "#2e6f9b"
systems: []
subsystems: []
containers: []
components: []
code: []
users: []
interfaces: []
relationships: []
"""
    source_path = tmp_path / "source.yaml"
    source_path.write_text(theme_yaml, encoding="utf-8")
    workbook_path = tmp_path / "architecture.xlsx"
    output_path = tmp_path / "output.yaml"

    source = load_architecture(source_path)
    write_workbook(source, workbook_path)
    from_excel = read_workbook(workbook_path)
    dump_architecture(from_excel, output_path)

    assert from_excel == source
    assert load_architecture(output_path) == source
    output_yaml = output_path.read_text(encoding="utf-8")
    assert output_yaml.index("timelines:") < output_yaml.index(
        "theme:"
    ) < output_yaml.index("systems:")
    assert build_payload(source, source_path.name)["theme"] == {
        "kinds": {"system": "#2e6f9b"}
    }

    invalid_path = tmp_path / "invalid.yaml"
    invalid_path.write_text(theme_yaml.replace("#2e6f9b", "teal"), encoding="utf-8")
    invalid = next(
        finding
        for finding in validate(load_architecture(invalid_path))
        if finding.code == "invalid_color"
    )
    assert (invalid.file, invalid.path, invalid.line > 1, invalid.column > 0) == (
        str(invalid_path),
        "theme.kinds.system",
        True,
        True,
    )

    unknown_path = tmp_path / "unknown.yaml"
    unknown_path.write_text(
        theme_yaml.replace('    system: "#2e6f9b"', '    database: "#2e6f9b"'),
        encoding="utf-8",
    )
    unknown = next(
        finding
        for finding in validate(load_architecture(unknown_path))
        if finding.code == "unknown_theme_key"
    )
    assert unknown.path == "theme.kinds.database"
