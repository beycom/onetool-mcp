"""Unit tests for the chat_ops pack."""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError


@pytest.mark.unit
@pytest.mark.tools
class TestChatOpsPack:
    """Validate pack surface and core behaviors."""

    def test_pack_names_and_exports(self):
        from otdev.tools import chat_ops

        assert chat_ops.pack == "chat_ops"
        assert chat_ops.pack_aliases == ("co",)
        expected = {"ingest", "report_excel", "report_summary", "report_llm", "note", "rebuild"}
        assert expected == set(chat_ops.__all__)

    def test_default_analysis_config_has_canonical_categories_and_signals(self):
        from otdev.tools import chat_ops

        cfg = chat_ops.Config()
        categories = [rule.category for rule in cfg.analysis.categories]
        assert categories == [
            "PLANNING",
            "DELEGATION",
            "TESTING",
            "VERSION_CONTROL",
            "BUILD_AND_DEPLOYMENT",
            "DEBUGGING",
            "REFACTORING",
            "FEATURE_DEVELOPMENT",
            "BRAINSTORMING",
            "EXPLORATION",
            "CODING",
            "CONVERSATION",
            "GENERAL",
        ]
        assert [rule.signal_type for rule in cfg.analysis.signals] == [
            "FAILURE",
            "RETRY",
            "PATCH_SUMMARY",
        ]

    def test_apply_analysis_rules_uses_signal_rules(self, monkeypatch: pytest.MonkeyPatch):
        from otdev.tools import chat_ops

        captured: dict[str, object] = {}

        def _fake_configure_analysis_rules(**kwargs):
            captured.update(kwargs)

        monkeypatch.setattr(chat_ops, "configure_analysis_rules", _fake_configure_analysis_rules)
        chat_ops._apply_analysis_rules(chat_ops.Config())

        assert captured["default_category"] == "GENERAL"
        assert "signal_rules" in captured
        assert "category_rules" in captured

    def test_provider_config_rejects_legacy_parser_field(self, tmp_path):
        from otdev.tools import chat_ops

        with pytest.raises(ValidationError):
            chat_ops.ProviderConfig(
                provider_dir=str(tmp_path),
                parser="codex",
            )

    def test_note_writes_event_annotation_row(self, tmp_path, monkeypatch: pytest.MonkeyPatch):
        from otdev.tools import chat_ops

        db_path = tmp_path / "chat_ops.db"
        monkeypatch.setattr(chat_ops, "_resolve_db_path", lambda _cfg: db_path)

        result = chat_ops.note(type="summary", message="summary text", session_id="session-1")
        assert isinstance(result, dict)
        assert result["ok"] is True

        conn = sqlite3.connect(db_path)
        row = conn.execute(
            """
            SELECT ea.annotation_type, ea.annotation_value, ea.source
            FROM event_annotations ea
            JOIN events e ON e.event_id = ea.event_id
            ORDER BY ea.event_annotation_id DESC
            LIMIT 1
            """
        ).fetchone()
        conn.close()

        assert row == ("summary", "summary text", "tool_call")

    def test_note_rejects_invalid_type(self, tmp_path, monkeypatch: pytest.MonkeyPatch):
        from otdev.tools import chat_ops

        db_path = tmp_path / "chat_ops.db"
        monkeypatch.setattr(chat_ops, "_resolve_db_path", lambda _cfg: db_path)

        result = chat_ops.note(type="invalid", message="summary text")
        assert isinstance(result, str)
        assert "invalid type 'invalid'" in result

    def test_report_excel_exports_v3_tabs_with_canonical_columns(self, tmp_path, monkeypatch):
        from otdev.tools import chat_ops

        db_path = tmp_path / "chat_ops.db"
        monkeypatch.setattr(chat_ops, "_resolve_db_path", lambda _cfg: db_path)
        monkeypatch.setattr(chat_ops, "resolve_cwd_path", lambda p: tmp_path / p)

        note_result = chat_ops.note(type="note", message="hello", session_id="s1", turn_id="t1")
        assert isinstance(note_result, dict)
        event_id = note_result["event_id"]

        conn = sqlite3.connect(db_path)
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute(
            """
            UPDATE sessions
            SET project = ?, started_at = ?, updated_at = ?, provider_session_id = ?
            WHERE session_id = ?
            """,
            (
                "proj-a",
                "2026-05-01T10:00:00+00:00",
                "2026-05-01T10:05:00+00:00",
                "provider-s1",
                note_result["session_id"],
            ),
        )
        conn.execute(
            """
            UPDATE turns
            SET turn_kind = 'explicit', turn_id_source = 'event_payload', started_at = '2026-05-01T10:01:00+00:00'
            WHERE turn_id = ?
            """,
            (note_result["turn_id"],),
        )
        conn.execute(
            """
            INSERT INTO event_commands(
                event_id, raw_command, family, base_cmd, subcommand, status, exit_code, duration_ms
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (event_id, "rg -n foo src", "shell", "rg", "-n", "succeeded", 0, 15.0),
        )
        conn.execute(
            """
            INSERT INTO event_usage(
                event_id, model, input_tokens, output_tokens, cached_input_tokens,
                reasoning_tokens, uncached_input_tokens, prompt_tokens, completion_tokens, total_tokens
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (event_id, "gpt-5", 10, 5, 0, 0, 10, 10, 5, 15),
        )
        conn.execute(
            """
            INSERT INTO mv_turn_metrics(
                turn_id, session_id, turn_kind, turn_started_at, model, project, task_category,
                commands_count, failed_commands_count, successful_commands_count,
                retry_count, one_shot_success,
                file_reads_count, files_read_unique, file_writes_count, files_written_unique,
                edit_churn, invocations_count, tool_calls_count,
                input_tokens, output_tokens, cached_input_tokens, reasoning_tokens,
                uncached_input_tokens, prompt_tokens, completion_tokens, total_tokens
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                note_result["turn_id"],
                note_result["session_id"],
                "explicit",
                "2026-05-01T10:01:00+00:00",
                "gpt-5",
                "proj-a",
                "CODING",
                1,
                0,
                1,
                0,
                1,
                0,
                0,
                0,
                0,
                0,
                0,
                0,
                10,
                5,
                0,
                0,
                10,
                10,
                5,
                15,
            ),
        )
        conn.execute(
            """
            INSERT INTO mv_session_metrics(
                session_id, provider, provider_session_id, rollout_id,
                project, session_name, first_user_message, started_at, updated_at,
                turns_count, commands_count, tool_calls_count, input_tokens, output_tokens, total_tokens
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                note_result["session_id"],
                "co",
                "provider-s1",
                "co.note",
                "proj-a",
                "Session A",
                "hello",
                "2026-05-01T10:00:00+00:00",
                "2026-05-01T10:05:00+00:00",
                1,
                1,
                0,
                10,
                5,
                15,
            ),
        )
        conn.commit()
        conn.close()

        result = chat_ops.report_excel(
            projects="proj-a",
            session_ids=note_result["session_id"],
            models="gpt-5",
            start="2026-05-01T00:00:00+00:00",
            end="2026-05-02T00:00:00+00:00",
        )
        assert isinstance(result, dict)
        assert result["format"] == "xlsx"
        output_path = Path(result["output_path"])
        assert output_path.exists()

        wb = load_workbook(output_path)
        ws = wb["commands"]
        headers = [cell.value for cell in ws[1]]
        assert headers[:4] == ["project", "session_id", "model", "date"]
        row = [cell.value for cell in ws[2]]
        assert row[0] == "proj-a"
        assert row[2] == "gpt-5"

    def test_report_excel_rejects_invalid_mode(self, tmp_path, monkeypatch: pytest.MonkeyPatch):
        from otdev.tools import chat_ops

        db_path = tmp_path / "chat_ops.db"
        monkeypatch.setattr(chat_ops, "_resolve_db_path", lambda _cfg: db_path)
        note_result = chat_ops.note(type="note", message="hello", session_id="s1")
        assert isinstance(note_result, dict)

        result = chat_ops.report_excel(report=["kpi"])
        assert isinstance(result, str)
        assert "unsupported excel report mode 'kpi'" in result

    def test_report_excel_rejects_non_list_report_arg(self, tmp_path, monkeypatch: pytest.MonkeyPatch):
        from otdev.tools import chat_ops

        db_path = tmp_path / "chat_ops.db"
        monkeypatch.setattr(chat_ops, "_resolve_db_path", lambda _cfg: db_path)
        note_result = chat_ops.note(type="note", message="hello", session_id="s1")
        assert isinstance(note_result, dict)

        result = chat_ops.report_excel(report="commands")  # type: ignore[arg-type]
        assert isinstance(result, str)
        assert "report must be a list[str] when provided" in result

    def test_report_summary_writes_yaml(self, tmp_path, monkeypatch: pytest.MonkeyPatch):
        import yaml

        from otdev.tools import chat_ops

        db_path = tmp_path / "chat_ops.db"
        monkeypatch.setattr(chat_ops, "_resolve_db_path", lambda _cfg: db_path)
        monkeypatch.setattr(chat_ops, "resolve_cwd_path", lambda p: tmp_path / p)
        note_result = chat_ops.note(type="note", message="implement summary report", session_id="s1")
        assert isinstance(note_result, dict)

        result = chat_ops.report_summary(output_name="summary.yaml")
        assert isinstance(result, dict)
        assert result["format"] == "yaml"
        out = Path(result["output_path"])
        assert out.exists()
        data = yaml.safe_load(out.read_text(encoding="utf-8"))
        assert "sessions" in data
        assert data["sessions"][0]["session_id"] == "s1"

    def test_report_llm_requires_model(self, tmp_path, monkeypatch: pytest.MonkeyPatch):
        from otdev.tools import chat_ops

        db_path = tmp_path / "chat_ops.db"
        monkeypatch.setattr(chat_ops, "_resolve_db_path", lambda _cfg: db_path)
        monkeypatch.setattr(chat_ops, "resolve_cwd_path", lambda p: tmp_path / p)
        note_result = chat_ops.note(type="note", message="hello", session_id="s1")
        assert isinstance(note_result, dict)

        result = chat_ops.report_llm()
        assert isinstance(result, str)
        assert "LLM summaries required but llm_model is not configured" in result

    def test_rebuild_uses_busy_timeout_on_sqlite_connection(self, tmp_path, monkeypatch):
        from otdev.tools import chat_ops

        db_path = tmp_path / "chat_ops.db"
        db_path.write_text("")
        monkeypatch.setattr(chat_ops, "_resolve_db_path", lambda _cfg: db_path)
        monkeypatch.setattr(chat_ops, "_apply_analysis_rules", lambda _cfg: None)

        captured: dict[str, object] = {}

        class _FakeConn:
            def __init__(self):
                self.executed: list[str] = []

            def execute(self, sql: str):
                self.executed.append(sql)
                return None

            def commit(self):
                return None

            def close(self):
                return None

        fake_conn = _FakeConn()

        def _fake_connect(path, timeout):
            captured["path"] = path
            captured["timeout"] = timeout
            return fake_conn

        monkeypatch.setattr(chat_ops.sqlite3, "connect", _fake_connect)
        monkeypatch.setattr(chat_ops, "rebuild_projections", lambda _conn: None)

        result = chat_ops.rebuild(provider="codex")
        assert isinstance(result, dict)
        assert result["ok"] is True
        assert captured["path"] == db_path
        assert captured["timeout"] == chat_ops._SQLITE_TIMEOUT_S
        assert f"PRAGMA busy_timeout={chat_ops._SQLITE_BUSY_TIMEOUT_MS}" in fake_conn.executed
        assert "PRAGMA journal_mode=WAL" in fake_conn.executed
        assert "PRAGMA foreign_keys=ON" in fake_conn.executed

    def test_pipeline_ingest_uses_consistent_sqlite_pragmas(self, tmp_path, monkeypatch):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()

        class _FakeCursor:
            lastrowid = 1

        class _FakeConn:
            def __init__(self):
                self.executed: list[str] = []

            def execute(self, sql: str, *_args):
                self.executed.append(sql)
                if "INSERT INTO ingest_runs" in sql:
                    return _FakeCursor()
                return None

            def executescript(self, sql: str):
                self.executed.append(sql)
                return None

            def commit(self):
                return None

            def rollback(self):
                return None

            def close(self):
                return None

        captured: dict[str, object] = {}
        fake_conn = _FakeConn()

        def _fake_connect(path, timeout):
            captured["path"] = path
            captured["timeout"] = timeout
            return fake_conn

        monkeypatch.setattr(pipeline.sqlite3, "connect", _fake_connect)
        monkeypatch.setattr(pipeline, "_discover_rollouts", lambda _dir: [])

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda *_args: None,
            )
        )

        assert result.scanned == 0
        assert captured["path"] == db_path
        assert captured["timeout"] == pipeline._SQLITE_TIMEOUT_S
        assert f"PRAGMA busy_timeout={pipeline._SQLITE_BUSY_TIMEOUT_MS}" in fake_conn.executed
        assert "PRAGMA journal_mode=WAL" in fake_conn.executed
        assert "PRAGMA foreign_keys=ON" in fake_conn.executed

    def test_ingest_uses_configured_providers_without_db_arg(self, tmp_path, monkeypatch):
        from otdev.tools import chat_ops

        p1 = tmp_path / "codex"
        p2 = tmp_path / "claude"
        p1.mkdir()
        p2.mkdir()
        db_path = tmp_path / "chat_ops.db"

        cfg = chat_ops.Config(
            providers={
                "codex": chat_ops.ProviderConfig(
                    provider_dir=str(p1),
                    parser_file="builtin:codex_parser",
                ),
                "claude": chat_ops.ProviderConfig(
                    provider_dir=str(p2),
                    parser_file="builtin:codex_parser",
                ),
            }
        )

        monkeypatch.setattr(chat_ops, "_get_config", lambda: cfg)
        monkeypatch.setattr(chat_ops, "_resolve_db_path", lambda _cfg: db_path)
        monkeypatch.setattr(chat_ops, "_apply_analysis_rules", lambda _cfg: None)

        class _Result:
            def __init__(self, run_id: int):
                self.run_id = run_id
                self.scanned = 1
                self.inserted = 2
                self.duplicates = 0
                self.skipped = 0
                self.errors = 0

        run_ids = {"codex": 10, "claude": 20}

        def _fake_pipeline_ingest(options):
            return _Result(run_ids[options.provider])

        monkeypatch.setattr(chat_ops, "pipeline_ingest", _fake_pipeline_ingest)

        result = chat_ops.ingest()
        assert isinstance(result, dict)
        providers = result["providers"]
        assert [item["provider"] for item in providers] == ["codex", "claude"]

    def test_pipeline_v3_schema_indexes_and_fk(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        conn = sqlite3.connect(db_path)
        conn.executescript(pipeline._SCHEMA_SQL)

        events_indexes = {
            row[1]
            for row in conn.execute("PRAGMA index_list('events')").fetchall()
        }
        assert "idx_events_provider_source_loc" in events_indexes
        assert "idx_events_scope" in events_indexes

        fk_events = conn.execute("PRAGMA foreign_key_list('events')").fetchall()
        fk_cols = {(row[3], row[4], row[2]) for row in fk_events}
        assert ("session_id", "session_id", "sessions") in fk_cols
        assert ("turn_id", "turn_id", "turns") in fk_cols
        conn.close()

    def test_pipeline_turn_assignment_precedence_and_usage_sparsity(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-05T10-00-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            "\n".join(
                [
                    json.dumps(
                        {
                            "type": "turn_context",
                            "session_id": "provider-session-1",
                            "turn_id": "ctx-turn-1",
                            "timestamp": "2026-05-05T10:00:00Z",
                        }
                    ),
                    json.dumps(
                        {
                            "type": "event_msg",
                            "timestamp": "2026-05-05T10:00:01Z",
                            "payload": {
                                "type": "exec_command_end",
                                "turn_id": "payload-turn-2",
                                "command": "echo hi",
                                "exit_code": 0,
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "response_item",
                            "timestamp": "2026-05-05T10:00:02Z",
                            "payload": {
                                "type": "message",
                                "role": "assistant",
                                "content": [{"text": "no turn id, no usage"}],
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "response_item",
                            "timestamp": "2026-05-05T10:00:03Z",
                            "payload": {
                                "type": "message",
                                "role": "assistant",
                                "content": [{"text": "has usage"}],
                                "usage": {
                                    "model": "gpt-5",
                                    "input_tokens": 12,
                                    "output_tokens": 4,
                                    "cached_input_tokens": 2,
                                    "reasoning_output_tokens": 1,
                                },
                            },
                        }
                    ),
                ]
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=True,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 4

        conn = sqlite3.connect(db_path)
        sessions = conn.execute(
            "SELECT provider_session_id FROM sessions LIMIT 1"
        ).fetchone()
        assert sessions is not None
        assert sessions[0] == "provider-session-1"

        sources = {
            row[0]
            for row in conn.execute("SELECT turn_id_source FROM turns").fetchall()
        }
        assert "turn_context" in sources
        assert "event_payload" in sources

        usage_count = conn.execute("SELECT COUNT(*) FROM event_usage").fetchone()[0]
        assert usage_count == 1

        projection_count = conn.execute("SELECT COUNT(*) FROM mv_turn_metrics").fetchone()[0]
        assert projection_count >= 1
        conn.close()

    def test_pipeline_resets_active_turn_when_session_changes_mid_stream(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-06T12-00-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            "\n".join(
                [
                    json.dumps(
                        {
                            "type": "turn_context",
                            "session_id": "provider-session-a",
                            "turn_id": "turn-a",
                            "timestamp": "2026-05-06T12:00:00Z",
                        }
                    ),
                    json.dumps(
                        {
                            "type": "response_item",
                            "timestamp": "2026-05-06T12:00:01Z",
                            "payload": {
                                "type": "message",
                                "role": "assistant",
                                "content": [{"text": "still session a"}],
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "event_msg",
                            "timestamp": "2026-05-06T12:00:02Z",
                            "payload": {
                                "type": "thread_name_updated",
                                "thread_id": "provider-session-b",
                                "thread_name": "session b",
                            },
                        }
                    ),
                ]
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 3

        conn = sqlite3.connect(db_path)
        row = conn.execute(
            """
            SELECT e.session_id, t.session_id
            FROM events e
            JOIN turns t ON t.turn_id = e.turn_id
            WHERE e.payload_type = 'thread_name_updated'
            LIMIT 1
            """
        ).fetchone()
        assert row is not None
        assert row[0] == row[1]
        conn.close()

    def test_pipeline_ignores_active_turn_when_it_belongs_to_other_session(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        conn = sqlite3.connect(db_path)
        conn.execute("PRAGMA foreign_keys=ON")
        conn.executescript(pipeline._SCHEMA_SQL)
        conn.execute(
            """
            INSERT INTO sessions(
                session_id, provider, rollout_id
            ) VALUES (?, ?, ?), (?, ?, ?)
            """,
            ("session-a", "codex", "rollout-a", "session-b", "codex", "rollout-b"),
        )
        conn.execute(
            """
            INSERT INTO turns(
                turn_id, session_id, provider_turn_id, turn_kind, turn_id_source
            ) VALUES (?, ?, ?, ?, ?)
            """,
            ("turn-a", "session-a", "provider-turn-a", "explicit", "event_payload"),
        )

        result = pipeline._resolve_turn_for_event(
            conn,
            payload={"type": "response_item", "payload": {"type": "message", "role": "assistant"}},
            session_id="session-b",
            active_turn_id="turn-a",
            event_seed="seed-1",
            event_ts="2026-05-06T12:00:03Z",
        )

        assert result["source"] == "ingest_synthetic"
        turn_row = conn.execute(
            "SELECT session_id FROM turns WHERE turn_id = ?",
            (result["turn_id"],),
        ).fetchone()
        assert turn_row is not None
        assert turn_row[0] == "session-b"
        conn.close()

    def test_pipeline_extracts_nested_session_meta_and_turn_context_payload_fields(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-06T13-00-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            "\n".join(
                [
                    json.dumps(
                        {
                            "type": "session_meta",
                            "timestamp": "2026-05-06T13:00:00Z",
                            "payload": {
                                "id": "provider-thread-123",
                                "originator": "codex",
                                "cli_version": "0.99.0",
                                "cwd": "/repo",
                                "model_provider": "openai",
                                "git_sha": "abc123",
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "turn_context",
                            "timestamp": "2026-05-06T13:00:01Z",
                            "payload": {
                                "turn_id": "turn-ctx-1",
                                "cwd": "/repo",
                                "model": "gpt-5.3-codex",
                                "approval_policy": "never",
                                "sandbox_policy": {"type": "danger-full-access"},
                                "collaboration_mode": {"mode": "default"},
                                "effort": "high",
                            },
                        }
                    ),
                ]
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 2

        conn = sqlite3.connect(db_path)
        meta = conn.execute(
            """
            SELECT provider_thread_id, originator, cli_version, cwd, model_provider, git_sha
            FROM event_session_meta
            LIMIT 1
            """
        ).fetchone()
        assert meta == ("provider-thread-123", "codex", "0.99.0", "/repo", "openai", "abc123")

        turn_context = conn.execute(
            """
            SELECT provider_turn_id, cwd, model, approval_policy, sandbox_policy, reasoning_effort
            FROM event_turn_context
            LIMIT 1
            """
        ).fetchone()
        assert turn_context is not None
        assert turn_context[0] == "turn-ctx-1"
        assert turn_context[1] == "/repo"
        assert turn_context[2] == "gpt-5.3-codex"
        assert turn_context[3] == "never"
        assert '"type": "danger-full-access"' in (turn_context[4] or "")
        assert turn_context[5] == "high"

        provider_session = conn.execute(
            "SELECT provider_session_id FROM sessions LIMIT 1"
        ).fetchone()
        assert provider_session is not None
        assert provider_session[0] == "provider-thread-123"
        conn.close()

    def test_pipeline_extracts_codex_token_count_usage_and_rate_limit_shapes(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-06T13-10-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            json.dumps(
                {
                    "type": "event_msg",
                    "timestamp": "2026-05-06T13:10:00Z",
                    "payload": {
                        "type": "token_count",
                        "info": {
                            "model": "gpt-5.3-codex",
                            "model_context_window": 258400,
                            "total_token_usage": {
                                "input_tokens": 1200,
                                "cached_input_tokens": 200,
                                "output_tokens": 50,
                                "reasoning_output_tokens": 10,
                                "total_tokens": 1250,
                            },
                            "last_token_usage": {
                                "input_tokens": 300,
                                "cached_input_tokens": 64,
                                "output_tokens": 21,
                                "reasoning_output_tokens": 7,
                                "total_tokens": 321,
                            },
                        },
                        "rate_limits": {
                            "limit_id": "codex",
                            "limit_name": "GPT-5.3-Codex",
                            "primary": {"used_percent": 4.0, "window_minutes": 300, "resets_at": 12345},
                            "secondary": {"used_percent": 1.0, "window_minutes": 10080, "resets_at": 99999},
                            "credits": {"has_credits": False, "unlimited": False, "balance": None},
                            "plan_type": "prolite",
                            "rate_limit_reached_type": None,
                        },
                    },
                }
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 1

        conn = sqlite3.connect(db_path)
        usage = conn.execute(
            """
            SELECT
                model, input_tokens, output_tokens, cached_input_tokens, reasoning_tokens,
                model_context_window, last_input_tokens, last_total_tokens,
                rate_limit_id, rate_limit_name, rate_primary_used_percent, plan_type
            FROM event_usage
            LIMIT 1
            """
        ).fetchone()
        assert usage is not None
        assert usage[0] == "gpt-5.3-codex"
        assert usage[1] == 300
        assert usage[2] == 21
        assert usage[3] == 64
        assert usage[4] == 7
        assert usage[5] == 258400
        assert usage[6] == 300
        assert usage[7] == 321
        assert usage[8] == "codex"
        assert usage[9] == "GPT-5.3-Codex"
        assert usage[10] == 4.0
        assert usage[11] == "prolite"
        conn.close()

    def test_pipeline_correlates_function_call_output_and_dedupes_tool_result_edges(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-06T13-20-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            "\n".join(
                [
                    json.dumps(
                        {
                            "type": "response_item",
                            "timestamp": "2026-05-06T13:20:00Z",
                            "payload": {
                                "type": "function_call",
                                "name": "mcp__onetool__run",
                                "arguments": "{\"command\":\"ot.debug()\"}",
                                "call_id": "call-123",
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "response_item",
                            "timestamp": "2026-05-06T13:20:01Z",
                            "payload": {
                                "type": "function_call_output",
                                "call_id": "call-123",
                                "output": "ok",
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "event_msg",
                            "timestamp": "2026-05-06T13:20:02Z",
                            "payload": {
                                "type": "mcp_tool_call_end",
                                "call_id": "call-123",
                                "invocation": {
                                    "server": "onetool",
                                    "tool": "run",
                                    "arguments": {"command": "ot.debug()"},
                                },
                                "duration": {"secs": 0, "nanos": 10},
                                "result": {
                                    "Ok": {
                                        "content": [{"type": "text", "text": "done"}],
                                        "isError": False,
                                    }
                                },
                            },
                        }
                    ),
                ]
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 3

        conn = sqlite3.connect(db_path)
        row = conn.execute(
            """
            SELECT
                call_id,
                tool_name,
                server,
                status,
                duration_ms,
                arguments_json,
                result_json,
                assistant_event_id,
                result_event_id
            FROM event_tool_calls
            WHERE call_id = ?
            LIMIT 1
            """,
            ("call-123",),
        ).fetchone()
        assert row is not None
        assert row[0] == "call-123"
        assert row[1] == "mcp__onetool__run"
        assert row[2] == "onetool"
        assert row[3] == "succeeded"
        assert row[4] is not None
        assert float(row[4]) >= 2000.0
        assert row[5] is not None
        assert "ot.debug()" in row[5]
        assert row[6] is not None
        assert row[7] is not None
        assert row[8] is not None

        linked = conn.execute(
            """
            SELECT COUNT(*)
            FROM event_tool_calls
            WHERE call_id = ?
              AND assistant_event_id IS NOT NULL
              AND result_event_id IS NOT NULL
            """,
            ("call-123",),
        ).fetchone()
        assert linked is not None
        assert linked[0] == 1

        edges = conn.execute(
            "SELECT COUNT(*) FROM event_edges WHERE edge_type = 'tool_result_of'"
        ).fetchone()
        assert edges is not None
        assert edges[0] == 1

        turn_tool_calls = conn.execute(
            "SELECT COALESCE(SUM(tool_calls_count), 0) FROM mv_turn_metrics"
        ).fetchone()
        assert turn_tool_calls is not None
        assert turn_tool_calls[0] == 1
        conn.close()

    def test_pipeline_keeps_mcp_tool_call_end_rows_out_of_event_commands(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-06T13-30-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            json.dumps(
                {
                    "type": "event_msg",
                    "timestamp": "2026-05-06T13:30:00Z",
                    "payload": {
                        "type": "mcp_tool_call_end",
                        "call_id": "call-456",
                        "invocation": {
                            "server": "onetool",
                            "tool": "run",
                            "arguments": {"command": "ground.search(query='x')"},
                        },
                        "duration": {"secs": 2, "nanos": 0},
                        "result": {
                            "Ok": {
                                "content": [{"type": "text", "text": "search output"}],
                                "isError": False,
                            }
                        },
                    },
                }
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 1

        conn = sqlite3.connect(db_path)
        commands_count = conn.execute("SELECT COUNT(*) FROM event_commands").fetchone()
        assert commands_count is not None
        assert commands_count[0] == 0

        tool_row = conn.execute(
            """
            SELECT call_id, tool_name, status
            FROM event_tool_calls
            WHERE call_id = ?
            LIMIT 1
            """,
            ("call-456",),
        ).fetchone()
        assert tool_row is not None
        assert tool_row[0] == "call-456"
        assert tool_row[1] == "run"
        assert tool_row[2] == "succeeded"
        conn.close()

    def test_pipeline_invocations_keep_intent_rows_only(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-06T13-40-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            "\n".join(
                [
                    json.dumps(
                        {
                            "type": "event_msg",
                            "timestamp": "2026-05-06T13:40:00Z",
                            "payload": {
                                "type": "user_message",
                                "message": "/review this\n$skill-install now",
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "response_item",
                            "timestamp": "2026-05-06T13:40:01Z",
                            "payload": {
                                "type": "function_call",
                                "name": "mcp__onetool__run",
                                "call_id": "call-xyz",
                                "arguments": "{\"command\":\"ot.debug()\"}",
                            },
                        }
                    ),
                ]
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 2

        conn = sqlite3.connect(db_path)
        invocation_types = {
            row[0]
            for row in conn.execute("SELECT DISTINCT invocation_type FROM event_invocations").fetchall()
        }
        assert "slash" in invocation_types
        assert "dollar" in invocation_types
        assert "skill" not in invocation_types
        assert "tool" not in invocation_types
        assert "agent" not in invocation_types
        conn.close()

    def test_pipeline_synthesizes_shell_commands_and_excludes_exec_command_tool_calls(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-06T13-45-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            "\n".join(
                [
                    json.dumps(
                        {
                            "type": "response_item",
                            "timestamp": "2026-05-06T13:45:00Z",
                            "payload": {
                                "type": "function_call",
                                "name": "exec_command",
                                "call_id": "call-shell-1",
                                "arguments": "{\"cmd\":\"rg --files\"}",
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "response_item",
                            "timestamp": "2026-05-06T13:45:01Z",
                            "payload": {
                                "type": "function_call_output",
                                "call_id": "call-shell-1",
                                "output": "src/main.py",
                            },
                        }
                    ),
                ]
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 2

        conn = sqlite3.connect(db_path)
        tool_rows = conn.execute(
            "SELECT COUNT(*) FROM event_tool_calls WHERE call_id = ?",
            ("call-shell-1",),
        ).fetchone()
        assert tool_rows is not None
        assert tool_rows[0] == 0

        command_row = conn.execute(
            """
            SELECT call_id, raw_command, status, source
            FROM event_commands
            WHERE call_id = ?
            LIMIT 1
            """,
            ("call-shell-1",),
        ).fetchone()
        assert command_row is not None
        assert command_row[0] == "call-shell-1"
        assert command_row[1] == "rg --files"
        assert command_row[2] == "succeeded"
        assert command_row[3] == "tool_call_synth"
        conn.close()

    def test_pipeline_applies_default_signal_rules_when_unconfigured(self, tmp_path):
        from onetool.chat_ops import pipeline

        pipeline.configure_analysis_rules(default_category="GENERAL", category_rules=[], signal_rules=[])

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-06T13-50-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            json.dumps(
                {
                    "type": "response_item",
                    "timestamp": "2026-05-06T13:50:00Z",
                    "payload": {
                        "type": "message",
                        "role": "assistant",
                        "content": [{"text": "command failed with error and needs retry"}],
                    },
                }
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 1

        conn = sqlite3.connect(db_path)
        signals = conn.execute("SELECT signal_type FROM event_signals").fetchall()
        assert signals
        signal_types = {row[0] for row in signals}
        assert "FAILURE" in signal_types or "RETRY" in signal_types
        conn.close()

    def test_pipeline_extracts_event_source_kind_from_codex_shape(self, tmp_path):
        from onetool.chat_ops import pipeline

        db_path = tmp_path / "chat_ops.db"
        provider_dir = tmp_path / "provider"
        provider_dir.mkdir()
        rollout = provider_dir / "rollout-2026-05-06T14-00-00-019df680-165a-7391-a148-aab581fe42d3.jsonl"
        rollout.write_text(
            json.dumps(
                {
                    "type": "event_msg",
                    "timestamp": "2026-05-06T14:00:00Z",
                    "payload": {
                        "type": "agent_message",
                        "event_source": {"kind": "extended"},
                        "message": "hello",
                    },
                }
            )
            + "\n"
        )

        result = pipeline.ingest(
            pipeline.IngestOptions(
                db_path=db_path,
                provider_dir=provider_dir,
                provider="codex",
                since=None,
                rebuild_projections=False,
                force_rescan=False,
                parse_version=1,
                parse_line=lambda line, *_args: json.loads(line),
            )
        )
        assert result.inserted == 1

        conn = sqlite3.connect(db_path)
        row = conn.execute("SELECT event_source_kind FROM events LIMIT 1").fetchone()
        assert row is not None
        assert row[0] == "extended"
        conn.close()
