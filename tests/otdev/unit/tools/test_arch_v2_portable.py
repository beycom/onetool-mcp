"""Schema-v2 YAML/Excel parity, scalar, extension, and source tests."""

from __future__ import annotations

from typing import TYPE_CHECKING

import pytest
from openpyxl import Workbook, load_workbook

from otdev.tools._arch.v2.load import load_workspace
from otdev.tools._arch.v2.models import ArchitectureWorkspace, Presentation
from otdev.tools._arch.v2.portable import semantic_payload
from otdev.tools._arch.v2.write import write_workspace
from tests.otdev.arch_v2_fixtures import ARCH_V2_FIXTURES

if TYPE_CHECKING:
    from pathlib import Path

pytestmark = [pytest.mark.unit, pytest.mark.tools]

FIXTURES = ARCH_V2_FIXTURES


def _minimal_workbook() -> Workbook:
    workbook = Workbook()
    active = workbook.active
    assert active is not None
    workbook.remove(active)
    return workbook


def test_yaml_excel_normalize_identically() -> None:
    """yaml-excel-normalize-identically: paired domain formats match semantically."""
    yaml_workspace = load_workspace(FIXTURES / "arch-v2-canonical.yaml").workspace
    excel_workspace = load_workspace(FIXTURES / "arch-v2-canonical.xlsx").workspace

    assert semantic_payload(yaml_workspace) == semantic_payload(excel_workspace)


def test_title_is_filename_stem_unless_configured(tmp_path: Path) -> None:
    """Presentation title is runtime configuration, not authored workbook data."""
    source = tmp_path / "customer-payments.yaml"
    source.write_text("schema_version: 2\nstates:\n  - id: base\n", encoding="utf-8")

    inferred = load_workspace(source).workspace
    configured = load_workspace(
        source,
        presentation=Presentation(title="Payments landscape"),
    ).workspace

    assert inferred.presentation.title == "customer-payments"
    assert configured.presentation.title == "Payments landscape"
    assert "presentation" not in semantic_payload(inferred)
    assert "title" not in semantic_payload(inferred)


def test_year_id_and_list_values(tmp_path: Path) -> None:
    """year-id-and-list-values: numeric IDs and bracket lists normalize predictably."""
    path = tmp_path / "year.xlsx"
    workbook = _minimal_workbook()
    change = workbook.create_sheet("change")
    change.append(["id", "name", "related_products"])
    change.append([2027, "2027 delivery", "[wallet;payments]"])
    system = workbook.create_sheet("sys")
    system.append(["change", "id", "name"])
    system.append([None, "A", "System A"])
    workbook.save(path)
    workbook.close()

    loaded = load_workspace(path)

    assert loaded.workspace.changes[0].id == "2027"
    assert loaded.workspace.changes[0].related_products == ["wallet", "payments"]


def test_blank_unset_removal_roundtrip(tmp_path: Path) -> None:
    """blank-unset-removal-roundtrip: blank remains absent and unset/removal explicit."""
    source = tmp_path / "sparse.xlsx"
    workbook = _minimal_workbook()
    change = workbook.create_sheet("change")
    change.append(["id", "name"])
    change.append(["delivery", "Delivery"])
    system = workbook.create_sheet("sys")
    system.append(["change", "id", "name", "description", "unset", "change_type"])
    system.append([None, "A", "System A", "Original", None, None])
    system.append([None, "B", "System B", None, None, None])
    system.append(["delivery", "A", None, None, "[description]", None])
    system.append(["delivery", "B", None, None, None, "removed"])
    workbook.save(source)
    workbook.close()

    loaded = load_workspace(source).workspace
    yaml_path = tmp_path / "sparse.yaml"
    excel_path = tmp_path / "roundtrip.xlsx"
    write_workspace(path=yaml_path, workspace=loaded)
    write_workspace(path=excel_path, workspace=load_workspace(yaml_path).workspace)
    final = load_workspace(excel_path).workspace
    patches = final.changes[0].patches.systems

    assert patches[0].unset == ["description"]
    assert "description" not in patches[0].model_fields_set
    assert patches[1].change_type == "removed"
    assert semantic_payload(final) == semantic_payload(loaded)


def test_extension_field_roundtrip(tmp_path: Path) -> None:
    """extension-field-roundtrip: typed and structured extension values survive Excel."""
    workspace = ArchitectureWorkspace.model_validate(
        {
            "schema_version": 2,
            "states": [
                {
                    "id": "base",
                    "systems": [
                        {
                            "id": "A",
                            "name": "A",
                            "portfolio": {"tier": 1, "owner": "platform"},
                        }
                    ],
                }
            ],
            "changes": [
                {
                    "id": "delivery",
                    "name": "Delivery",
                    "portfolio_code": "P-1",
                    "patches": {
                        "systems": [
                            {
                                "id": "A",
                                "risk": {"rating": "low"},
                            }
                        ]
                    },
                }
            ],
        }
    )
    path = tmp_path / "extensions.xlsx"

    write_workspace(path=path, workspace=workspace)
    loaded = load_workspace(path).workspace

    assert semantic_payload(loaded) == semantic_payload(workspace)
    assert loaded.states[0].systems[0].model_extra == {
        "portfolio": {"owner": "platform", "tier": 1}
    }
    assert loaded.changes[0].patches.systems[0].model_extra == {
        "risk": {"rating": "low"}
    }


def test_state_metadata_uses_domain_rows(tmp_path: Path) -> None:
    """State metadata round-trips without an Excel model worksheet."""
    workspace = ArchitectureWorkspace.model_validate(
        {
            "schema_version": 2,
            "states": [
                {
                    "id": "base",
                    "name": "Baseline",
                    "description": "Current architecture",
                    "properties": {"owner": "architecture"},
                    "portfolio": "core",
                    "systems": [{"id": "A", "name": "System A"}],
                }
            ],
        }
    )
    path = tmp_path / "state-metadata.xlsx"

    write_workspace(path=path, workspace=workspace)
    loaded = load_workspace(path).workspace

    assert semantic_payload(loaded) == semantic_payload(workspace)
    workbook = load_workbook(path, read_only=True)
    try:
        assert "model" not in workbook.sheetnames
    finally:
        workbook.close()


def test_yaml_source_location() -> None:
    """yaml-source-location: every canonical value retains its data path."""
    loaded = load_workspace(FIXTURES / "arch-v2-canonical.yaml")
    location = loaded.sources["views[0].through"]

    assert location.kind == "yaml"
    assert location.path.endswith("arch-v2-canonical.yaml")
    assert location.yaml_path == "views[0].through"


def test_excel_source_location() -> None:
    """excel-source-location: interface fields retain workbook/sheet/row/column."""
    loaded = load_workspace(FIXTURES / "arch-v2-canonical.xlsx")
    provider = next(
        location
        for key, location in loaded.sources.items()
        if key.startswith("interface[") and key.endswith(".provider")
    )
    interface_source = loaded.workspace.states[0].interfaces[0].source

    assert provider.kind == "excel"
    assert provider.sheet == "interface"
    assert provider.row is not None
    assert provider.column is not None
    assert interface_source is not None
    assert provider in interface_source.generated_from
