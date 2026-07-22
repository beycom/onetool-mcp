"""End-to-end schema-v2 conversion, initialization, and bundle tests."""

from __future__ import annotations

import json
import zipfile
from typing import TYPE_CHECKING

import pytest
from openpyxl import load_workbook

from otdev.tools import arch
from otdev.tools._arch.v2.load import load_workspace
from otdev.tools._arch.v2.portable import semantic_payload
from tests.otdev.arch_v2_fixtures import ARCH_V2_FIXTURES

if TYPE_CHECKING:
    from pathlib import Path

pytestmark = [pytest.mark.integration, pytest.mark.tools]

FIXTURES = ARCH_V2_FIXTURES


def test_yaml_excel_yaml_golden(tmp_path: Path) -> None:
    """yaml-excel-yaml-golden: public conversion preserves canonical semantics."""
    source = FIXTURES / "arch-v2-canonical.yaml"
    excel = tmp_path / "converted.xlsx"
    final = tmp_path / "converted.yaml"

    assert arch.convert(input_path=str(source), output_path=str(excel))["ok"] is True
    assert arch.convert(input_path=str(excel), output_path=str(final))["ok"] is True

    assert semantic_payload(load_workspace(final).workspace) == semantic_payload(
        load_workspace(source).workspace
    )
    converted = load_workspace(final).workspace
    assert converted.states[0].systems[0].group == ["payments", "core-platform"]
    assert converted.states[0].systems[0].properties == {
        "owner": "payments",
        "tier": "one",
    }
    assert converted.changes[0].group == ["wave-one"]
    workbook = load_workbook(excel, read_only=True)
    try:
        assert "model" not in workbook.sheetnames
    finally:
        workbook.close()


def test_excel_yaml_excel_golden(tmp_path: Path) -> None:
    """excel-yaml-excel-golden: reverse public conversion is semantically stable."""
    source = FIXTURES / "arch-v2-canonical.xlsx"
    yaml_path = tmp_path / "converted.yaml"
    final = tmp_path / "converted.xlsx"

    assert (
        arch.convert(input_path=str(source), output_path=str(yaml_path))["ok"] is True
    )
    assert arch.convert(input_path=str(yaml_path), output_path=str(final))["ok"] is True

    assert semantic_payload(load_workspace(final).workspace) == semantic_payload(
        load_workspace(source).workspace
    )
    converted = load_workspace(final).workspace
    assert converted.states[0].systems[0].group == ["payments", "core-platform"]
    assert converted.changes[0].group == ["wave-one"]
    workbook = load_workbook(final, read_only=True)
    try:
        assert "model" not in workbook.sheetnames
    finally:
        workbook.close()
    repeated = tmp_path / "repeated.xlsx"
    assert (
        arch.convert(input_path=str(yaml_path), output_path=str(repeated))["ok"] is True
    )
    assert final.read_bytes() == repeated.read_bytes()


def test_init_canonical_fixture_no_drift(tmp_path: Path) -> None:
    """init-canonical-fixture-no-drift: initialized examples share production semantics."""
    output = tmp_path / "new solution"

    result = arch.init(output_path=str(output))

    assert result["ok"] is True
    fixture = semantic_payload(
        load_workspace(FIXTURES / "arch-v2-canonical.yaml").workspace
    )
    yaml_workspace = semantic_payload(
        load_workspace(output / "architecture.yaml").workspace
    )
    excel_workspace = semantic_payload(
        load_workspace(output / "architecture.xlsx").workspace
    )
    assert fixture == yaml_workspace == excel_workspace
    assert (output / "views" / "platform-delivery.c4").is_file()
    assert (output / "styles" / "clean.yaml").is_file()
    assert (output / "assets" / "icons").is_dir()
    assert (output / "assets" / "attachments").is_dir()


def test_bundle_offline_spaced_path(tmp_path: Path) -> None:
    """bundle-offline-spaced-path: archives are deterministic and manifest-scoped."""
    workspace = tmp_path / "workspace with spaces"
    assert arch.init(output_path=str(workspace))["ok"] is True
    generated = workspace / "generated" / "report.html"
    generated.parent.mkdir()
    generated.write_text("<!doctype html><title>Offline</title>", encoding="utf-8")
    (workspace / "manifest.json").write_text(
        json.dumps({"artifacts": [{"path": "generated/report.html"}]}),
        encoding="utf-8",
    )
    without_generated = tmp_path / "bundle without generated.zip"
    repeated = tmp_path / "bundle repeated.zip"
    with_generated = tmp_path / "bundle with generated.zip"

    first = arch.bundle(
        input_path=str(workspace),
        output_path=str(without_generated),
    )
    second = arch.bundle(
        input_path=str(workspace),
        output_path=str(repeated),
    )
    included = arch.bundle(
        input_path=str(workspace),
        output_path=str(with_generated),
        include_generated=True,
    )

    assert first["ok"] is second["ok"] is included["ok"] is True
    assert without_generated.read_bytes() == repeated.read_bytes()
    with zipfile.ZipFile(without_generated) as archive:
        names = archive.namelist()
        assert "architecture.yaml" in names
        assert "views/platform-delivery.c4" in names
        assert "styles/clean.yaml" in names
        assert "assets/icons/" in names
        assert "generated/report.html" not in names
    with zipfile.ZipFile(with_generated) as archive:
        assert "generated/report.html" in archive.namelist()
