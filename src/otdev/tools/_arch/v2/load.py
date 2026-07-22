"""Production YAML and compact Excel schema-v2 workspace loaders."""

from __future__ import annotations

import json
from collections import Counter
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from pathlib import Path

import yaml

from .models import ArchitectureWorkspace, LoadedWorkspace, Presentation, SourceLocation

YAML_SUFFIXES = {".yaml", ".yml"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
ENTITY_SHEETS = ("sys", "app", "cmp", "interface", "usr")
LIST_FIELDS = {
    "tags",
    "group",
    "related_products",
    "depends_on",
    "focus",
    "display_statuses",
    "systems",
    "changes",
    "unset",
}


class WorkspaceLoadError(ValueError):
    """Raised when a workspace cannot be parsed into schema-v2 input."""


def canonical_id(value: Any) -> str:
    """Normalize scalar identities, including Excel year-like numeric cells."""
    if isinstance(value, bool):
        raise WorkspaceLoadError("Boolean values are not valid identifiers")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    if not text:
        raise WorkspaceLoadError("Identifiers must not be blank")
    return text


def parse_list_cell(value: Any) -> Any:
    """Parse deterministic bracketed semicolon lists from Excel."""
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if len(stripped) < 2 or not (stripped.startswith("[") and stripped.endswith("]")):
        return stripped
    inner = stripped[1:-1]
    if not inner.strip():
        return []
    return [item.strip() for item in inner.split(";")]


def parse_properties_cell(value: Any) -> dict[str, Any]:
    """Parse JSON or compact name/value properties from an Excel cell."""
    if isinstance(value, dict):
        return value
    if not isinstance(value, str):
        raise WorkspaceLoadError(
            "Excel properties must be a JSON object or name:value list"
        )
    text = value.strip()
    if not text:
        return {}
    if text.startswith("{"):

        def unique_object(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
            properties: dict[str, Any] = {}
            for key, item in pairs:
                if key in properties:
                    raise WorkspaceLoadError(f"Duplicate property name {key!r}")
                properties[key] = item
            return properties

        try:
            parsed = json.loads(text, object_pairs_hook=unique_object)
        except WorkspaceLoadError:
            raise
        except json.JSONDecodeError as exc:
            raise WorkspaceLoadError(f"Invalid JSON properties: {exc}") from exc
        if not isinstance(parsed, dict):
            raise WorkspaceLoadError("Excel JSON properties must contain an object")
        return parsed
    entries = [entry.strip() for line in text.splitlines() for entry in line.split(";")]
    properties: dict[str, Any] = {}
    for entry in entries:
        if not entry:
            continue
        if ":" not in entry:
            raise WorkspaceLoadError(
                f"Invalid property entry {entry!r}; expected name:value"
            )
        name, raw_value = entry.split(":", maxsplit=1)
        key = name.strip()
        if not key:
            raise WorkspaceLoadError("Property names must not be blank")
        if key in properties:
            raise WorkspaceLoadError(f"Duplicate property name {key!r}")
        properties[key] = raw_value.strip()
    return properties


def _yaml_sources(
    *, value: Any, path: str, file_path: Path
) -> dict[str, SourceLocation]:
    sources: dict[str, SourceLocation] = {}

    def visit(item: Any, data_path: str) -> None:
        sources[data_path] = SourceLocation(
            kind="yaml",
            path=str(file_path),
            yaml_path=data_path,
        )
        if isinstance(item, dict):
            for key, child in item.items():
                visit(child, f"{data_path}.{key}" if data_path else str(key))
        elif isinstance(item, list):
            for index, child in enumerate(item):
                visit(child, f"{data_path}[{index}]")

    visit(value, path)
    return sources


def _configured_presentation(
    presentation: Presentation | None,
    *,
    default_roadmap: str | None,
    source_title: str,
) -> Presentation:
    configured = presentation or Presentation()
    updates: dict[str, Any] = {}
    if "title" not in configured.model_fields_set:
        updates["title"] = source_title
    if configured.default_roadmap is None and default_roadmap is not None:
        updates["default_roadmap"] = default_roadmap
    return configured.model_copy(update=updates) if updates else configured


def load_yaml_workspace(
    path: Path, *, presentation: Presentation | None = None
) -> LoadedWorkspace:
    """Load a schema-v2 YAML workspace and retain every data path."""
    try:
        raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    except (OSError, yaml.YAMLError) as exc:
        raise WorkspaceLoadError(
            f"Unable to read YAML workspace '{path}': {exc}"
        ) from exc
    if not isinstance(raw, dict):
        raise WorkspaceLoadError(f"YAML workspace '{path}' must contain a mapping")
    authored_presentation = {key for key in ("title", "presentation") if key in raw}
    if authored_presentation:
        names = ", ".join(sorted(authored_presentation))
        raise WorkspaceLoadError(
            f"Workspace {names} belongs in tools.arch presentation configuration"
        )
    default_roadmap = next(
        (
            str(item["id"])
            for item in raw.get("roadmaps", [])
            if item.get("id") is not None
        ),
        None,
    )
    raw["presentation"] = _configured_presentation(
        presentation,
        default_roadmap=default_roadmap,
        source_title=path.stem,
    ).model_dump(mode="json", exclude_none=True)
    sources = _yaml_sources(value=raw, path="", file_path=path)
    for state_index, state in enumerate(raw.get("states", [])):
        state_path = f"states[{state_index}]"
        state.setdefault(
            "source", sources[state_path].model_dump(mode="json", exclude_none=True)
        )
        for field in (
            "systems",
            "applications",
            "components",
            "interfaces",
            "users",
            "relationships",
        ):
            for entity_index, entity in enumerate(state.get(field, [])):
                entity_path = f"{state_path}.{field}[{entity_index}]"
                entity.setdefault(
                    "source",
                    sources[entity_path].model_dump(mode="json", exclude_none=True),
                )
    for change_index, change in enumerate(raw.get("changes", [])):
        change_path = f"changes[{change_index}]"
        change.setdefault(
            "source", sources[change_path].model_dump(mode="json", exclude_none=True)
        )
        for field, patches in change.get("patches", {}).items():
            for patch_index, patch in enumerate(patches):
                patch_path = f"{change_path}.patches.{field}[{patch_index}]"
                patch.setdefault(
                    "source",
                    sources[patch_path].model_dump(mode="json", exclude_none=True),
                )
    for roadmap_index, roadmap in enumerate(raw.get("roadmaps", [])):
        roadmap_path = f"roadmaps[{roadmap_index}]"
        roadmap.setdefault(
            "source", sources[roadmap_path].model_dump(mode="json", exclude_none=True)
        )
        for item_index, item in enumerate(roadmap.get("items", [])):
            item_path = f"{roadmap_path}.items[{item_index}]"
            item.setdefault(
                "source", sources[item_path].model_dump(mode="json", exclude_none=True)
            )
    for view_index, view in enumerate(raw.get("views", [])):
        view_path = f"views[{view_index}]"
        view.setdefault(
            "source", sources[view_path].model_dump(mode="json", exclude_none=True)
        )
    for diagram_index, diagram in enumerate(raw.get("diagrams", [])):
        diagram_path = f"diagrams[{diagram_index}]"
        diagram.setdefault(
            "source_location",
            sources[diagram_path].model_dump(mode="json", exclude_none=True),
        )
    workspace = ArchitectureWorkspace.model_validate(raw)
    return LoadedWorkspace(
        workspace=workspace,
        format="yaml",
        path=str(path),
        sources=sources,
    )


def _normalise_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return "_".join(
        part for part in "".join(ch if ch.isalnum() else " " for ch in text).split()
    )


def _read_sheet_rows(
    *, workbook: Any, workbook_path: Path, sheet_name: str
) -> tuple[list[dict[str, Any]], dict[str, SourceLocation]]:
    if sheet_name not in workbook.sheetnames:
        return [], {}
    worksheet = workbook[sheet_name]
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1), None)
    if header_cells is None:
        return [], {}
    headers = [_normalise_header(cell.value) for cell in header_cells]
    duplicates = sorted(
        header
        for header, count in Counter(item for item in headers if item).items()
        if count > 1
    )
    if duplicates:
        raise WorkspaceLoadError(
            f"Workbook '{workbook_path}' sheet '{sheet_name}' has duplicate columns: {duplicates}"
        )

    rows: list[dict[str, Any]] = []
    sources: dict[str, SourceLocation] = {}
    for row_number, cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
        row: dict[str, Any] = {}
        cell_sources: dict[str, SourceLocation] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            cell = cells[index]
            value = cell.value
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            parsed = value.strip() if isinstance(value, str) else value
            structured = {
                "default_selection",
                "expected",
                "presentation",
                "state_extensions",
                "state_properties",
                "style",
                "tables",
                "themes",
                "variants",
            }
            looks_like_json = isinstance(parsed, str) and parsed.startswith(("{", "[{"))
            if header == "properties":
                try:
                    parsed = parse_properties_cell(parsed)
                except WorkspaceLoadError as exc:
                    raise WorkspaceLoadError(
                        f"Workbook '{workbook_path}' sheet '{sheet_name}' row {row_number} "
                        f"column {cell.column_letter} contains invalid properties: {exc}"
                    ) from exc
            elif (header in structured or looks_like_json) and isinstance(parsed, str):
                try:
                    parsed = json.loads(parsed)
                except json.JSONDecodeError as exc:
                    raise WorkspaceLoadError(
                        f"Workbook '{workbook_path}' sheet '{sheet_name}' row {row_number} "
                        f"column {cell.column_letter} contains invalid JSON: {exc}"
                    ) from exc
            else:
                parsed = parse_list_cell(parsed)
            if header in {"id", "change", "roadmap", "base", "state"}:
                parsed = canonical_id(parsed)
            row[header] = parsed
            location = SourceLocation(
                kind="excel",
                path=str(workbook_path),
                workbook=str(workbook_path),
                sheet=sheet_name,
                row=row_number,
                column=cell.column_letter,
            )
            sources[f"{sheet_name}[{row_number}].{header}"] = location
            cell_sources[header] = location
        if row:
            row["__row__"] = row_number
            row["__sources__"] = cell_sources
            rows.append(row)
    return rows, sources


def _source_for_row(
    *, path: Path, sheet: str, row: int, cells: dict[str, SourceLocation] | None = None
) -> dict[str, Any]:
    return SourceLocation(
        kind="excel",
        path=str(path),
        workbook=str(path),
        sheet=sheet,
        row=row,
        generated_from=list((cells or {}).values()),
    ).model_dump(mode="json", exclude_none=True)


def _entity_from_row(*, sheet: str, row: dict[str, Any], path: Path) -> dict[str, Any]:
    item = {
        key: value
        for key, value in row.items()
        if not key.startswith("__")
        and not key.startswith("state_")
        and key not in {"change", "entity_kind", "state"}
    }
    item["id"] = canonical_id(item["id"])
    item["source"] = _source_for_row(
        path=path,
        sheet=sheet,
        row=row["__row__"],
        cells=row["__sources__"],
    )
    return item


def _patch_from_row(*, sheet: str, row: dict[str, Any], path: Path) -> dict[str, Any]:
    item = _entity_from_row(sheet=sheet, row=row, path=path)
    item["id"] = canonical_id(item["id"])
    return item


def _selection_from_row(row: dict[str, Any]) -> dict[str, Any]:
    selection = {key: value for key, value in row.items() if not key.startswith("__")}
    for field in LIST_FIELDS:
        value = selection.get(field)
        if isinstance(value, str) and field in {"focus", "display_statuses"}:
            selection[field] = [value]
    return selection


def load_excel_workspace(
    path: Path, *, presentation: Presentation | None = None
) -> LoadedWorkspace:
    """Load the compact schema-v2 Excel mapping with blank-change base rows."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - optional dependency guard
        raise WorkspaceLoadError(
            "openpyxl is required for Excel architecture workspaces"
        ) from exc

    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
        sources: dict[str, SourceLocation] = {}
        for sheet in ("change", "roadmap", "view", "diagram", *ENTITY_SHEETS):
            rows, sheet_sources = _read_sheet_rows(
                workbook=workbook,
                workbook_path=path,
                sheet_name=sheet,
            )
            rows_by_sheet[sheet] = rows
            sources.update(sheet_sources)
    finally:
        workbook.close()

    explicit_bases = {
        canonical_id(row["base"])
        for row in rows_by_sheet["roadmap"]
        if row.get("base") is not None
    }
    entity_states = {
        canonical_id(row["state"])
        for sheet in ENTITY_SHEETS
        for row in rows_by_sheet[sheet]
        if row.get("change") is None and row.get("state") is not None
    }
    inferred_states = explicit_bases | entity_states
    if len(inferred_states) > 1:
        raise WorkspaceLoadError(
            f"Excel domain sheets must identify one base state; received {sorted(inferred_states)}"
        )
    state_id = next(iter(inferred_states), "base")
    base_rows = [
        row
        for sheet in ENTITY_SHEETS
        for row in rows_by_sheet[sheet]
        if row.get("change") is None
    ]

    def state_metadata(field: str) -> Any:
        values = [row[field] for row in base_rows if row.get(field) is not None]
        unique = {json.dumps(value, sort_keys=True, default=str) for value in values}
        if len(unique) > 1:
            raise WorkspaceLoadError(
                f"Excel domain sheets contain conflicting {field} values"
            )
        return values[0] if values else None

    generated_from = [
        source
        for sheet in ENTITY_SHEETS
        for row in rows_by_sheet[sheet]
        if row.get("change") is None
        for source in row["__sources__"].values()
    ]
    state: dict[str, Any] = {
        "id": state_id,
        "systems": [],
        "applications": [],
        "components": [],
        "interfaces": [],
        "users": [],
        "relationships": [],
        "source": SourceLocation(
            kind="generated",
            path=str(path),
            generated_from=generated_from,
        ).model_dump(mode="json", exclude_none=True),
    }
    metadata_fields = {
        "name": state_metadata("state_name"),
        "description": state_metadata("state_description"),
        "properties": state_metadata("state_properties"),
    }
    state.update(
        {key: value for key, value in metadata_fields.items() if value is not None}
    )
    extensions = state_metadata("state_extensions")
    if isinstance(extensions, dict):
        state.update(extensions)
    state_field = {
        "sys": "systems",
        "app": "applications",
        "cmp": "components",
        "interface": "interfaces",
        "usr": "users",
    }
    patches_by_change: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for sheet in ENTITY_SHEETS:
        for row in rows_by_sheet[sheet]:
            change_id = row.get("change")
            field = (
                "relationships"
                if sheet == "interface" and row.get("entity_kind") == "relationship"
                else state_field[sheet]
            )
            if change_id is None:
                state[field].append(_entity_from_row(sheet=sheet, row=row, path=path))
                continue
            change_key = canonical_id(change_id)
            groups = patches_by_change.setdefault(
                change_key,
                {
                    "systems": [],
                    "applications": [],
                    "components": [],
                    "interfaces": [],
                    "users": [],
                    "relationships": [],
                },
            )
            groups[field].append(_patch_from_row(sheet=sheet, row=row, path=path))

    changes: list[dict[str, Any]] = []
    for row in rows_by_sheet["change"]:
        change_id = canonical_id(row["id"])
        change = {key: value for key, value in row.items() if not key.startswith("__")}
        change["id"] = change_id
        change.setdefault("name", change_id)
        change["patches"] = patches_by_change.pop(
            change_id,
            {
                "systems": [],
                "applications": [],
                "components": [],
                "interfaces": [],
                "users": [],
                "relationships": [],
            },
        )
        change["source"] = _source_for_row(
            path=path,
            sheet="change",
            row=row["__row__"],
            cells=row["__sources__"],
        )
        changes.append(change)
    if patches_by_change:
        unknown = ", ".join(sorted(patches_by_change))
        raise WorkspaceLoadError(f"Entity patches reference unknown changes: {unknown}")

    roadmap_groups: dict[str, dict[str, Any]] = {}
    for row in rows_by_sheet["roadmap"]:
        if row.get("order") is None:
            raise WorkspaceLoadError(
                f"Excel roadmap sheet row {row['__row__']} requires an order"
            )
        roadmap_id = canonical_id(row.get("roadmap", "default"))
        roadmap = roadmap_groups.setdefault(
            roadmap_id,
            {
                "id": roadmap_id,
                "name": row.get("roadmap_name"),
                "base": canonical_id(row.get("base", state_id)),
                "items": [],
                "source": _source_for_row(
                    path=path,
                    sheet="roadmap",
                    row=row["__row__"],
                    cells=row["__sources__"],
                ),
            },
        )
        for key, value in row.items():
            if not key.startswith("__") and key not in {
                "roadmap",
                "roadmap_name",
                "base",
                "change",
                "order",
            }:
                roadmap.setdefault(key, value)
        roadmap["items"].append(
            {
                "change": canonical_id(row["change"]),
                "order": int(row["order"]),
                "source": _source_for_row(
                    path=path,
                    sheet="roadmap",
                    row=row["__row__"],
                    cells=row["__sources__"],
                ),
            }
        )

    views = []
    for row in rows_by_sheet["view"]:
        view = _selection_from_row(row)
        view["source"] = _source_for_row(
            path=path,
            sheet="view",
            row=row["__row__"],
            cells=row["__sources__"],
        )
        views.append(view)
    diagrams = []
    for row in rows_by_sheet["diagram"]:
        diagram = _selection_from_row(row)
        diagram["source_location"] = _source_for_row(
            path=path,
            sheet="diagram",
            row=row["__row__"],
            cells=row["__sources__"],
        )
        diagrams.append(diagram)
    default_roadmap = next(iter(roadmap_groups), None)
    configured_presentation = _configured_presentation(
        presentation,
        default_roadmap=default_roadmap,
        source_title=path.stem,
    )
    raw_workspace = {
        "schema_version": 2,
        "states": [state],
        "changes": changes,
        "roadmaps": list(roadmap_groups.values()),
        "views": views,
        "diagrams": diagrams,
        "presentation": configured_presentation.model_dump(
            mode="json", exclude_none=True
        ),
    }
    return LoadedWorkspace(
        workspace=ArchitectureWorkspace.model_validate(raw_workspace),
        format="excel",
        path=str(path),
        sources=sources,
    )


def load_workspace(
    path: Path, *, presentation: Presentation | None = None
) -> LoadedWorkspace:
    """Load a YAML or Excel schema-v2 workspace from a resolved path."""
    resolved = path.resolve()
    suffix = resolved.suffix.lower()
    if suffix in YAML_SUFFIXES:
        return load_yaml_workspace(resolved, presentation=presentation)
    if suffix in EXCEL_SUFFIXES:
        return load_excel_workspace(resolved, presentation=presentation)
    raise WorkspaceLoadError(f"Unsupported architecture workspace format: {resolved}")
