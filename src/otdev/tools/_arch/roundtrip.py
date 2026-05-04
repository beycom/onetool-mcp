"""Excel <-> YAML round-trip helpers for arch pack."""

from __future__ import annotations

from copy import copy
from typing import TYPE_CHECKING, Any

import yaml

from .models import SHEETS, normalize_key

if TYPE_CHECKING:
    from pathlib import Path


class RoundtripError(ValueError):
    """Raised for round-trip conversion errors."""


def _ensure_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - import guard
        raise ImportError(
            "openpyxl is required for arch pack. Install with: pip install onetool-mcp[dev]"
        ) from exc
    return openpyxl


def export_entities_to_yaml(*, entities: dict[str, list[dict[str, Any]]], output_path: Path) -> str:
    """Write entities to YAML file."""
    cleaned: dict[str, list[dict[str, Any]]] = {}
    for sheet in SHEETS:
        cleaned[sheet] = []
        for row in entities.get(sheet, []):
            cleaned[sheet].append({k: v for k, v in row.items() if not k.startswith("_")})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        yaml.safe_dump(cleaned, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )
    return str(output_path)


def load_yaml_entities(*, input_path: Path) -> dict[str, list[dict[str, Any]]]:
    """Load YAML entity sections."""
    if not input_path.exists():
        raise RoundtripError(f"YAML input not found: {input_path}")
    loaded = yaml.safe_load(input_path.read_text(encoding="utf-8"))
    if not isinstance(loaded, dict):
        raise RoundtripError("YAML root must be a mapping of sheet -> rows")

    entities: dict[str, list[dict[str, Any]]] = {}
    for sheet in SHEETS:
        rows = loaded.get(sheet, [])
        if not isinstance(rows, list):
            raise RoundtripError(f"YAML section '{sheet}' must be a list")
        parsed: list[dict[str, Any]] = []
        for row in rows:
            if not isinstance(row, dict):
                raise RoundtripError(f"YAML section '{sheet}' contains non-mapping row")
            parsed.append(dict(row))
        entities[sheet] = parsed
    return entities


def _copy_row_style(*, worksheet: Any, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = worksheet.cell(row=source_row, column=col)
        target = worksheet.cell(row=target_row, column=col)
        target._style = copy(source._style)
        if source.has_style:
            target.number_format = source.number_format
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)


def import_yaml_into_template(*, entities: dict[str, list[dict[str, Any]]], template_path: Path, output_path: Path) -> str:
    """Import YAML entities into workbook template and save new workbook."""
    openpyxl = _ensure_openpyxl()
    if not template_path.exists():
        raise RoundtripError(f"Template workbook not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    try:
        for sheet in SHEETS:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            header_cells = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=False)))
            headers = [normalize_key(cell.value) for cell in header_cells]
            max_col = len(headers)

            # Clear existing row values while preserving style/layout.
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).value = None

            rows = entities.get(sheet, [])
            style_source_row = 2 if ws.max_row >= 2 else 1
            for offset, row in enumerate(rows, start=2):
                if offset > ws.max_row:
                    ws.insert_rows(offset)
                    _copy_row_style(
                        worksheet=ws,
                        source_row=style_source_row,
                        target_row=offset,
                        max_col=max_col,
                    )
                for col_idx, header in enumerate(headers, start=1):
                    if not header:
                        continue
                    ws.cell(row=offset, column=col_idx).value = row.get(header)

            # Adjust table range to cover written rows.
            last_row = max(2, len(rows) + 1)
            for table in ws.tables.values():
                start_cell = table.ref.split(":", maxsplit=1)[0]
                col_end = header_cells[-1].column_letter if header_cells else "A"
                table.ref = f"{start_cell}:{col_end}{last_row}"

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    finally:
        wb.close()

    return str(output_path)


__all__ = [
    "RoundtripError",
    "export_entities_to_yaml",
    "import_yaml_into_template",
    "load_yaml_entities",
]
