"""Workbook discovery and ingestion for arch pack."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from otpack import resolve_cwd_path

from .models import SHEETS, normalize_cell, normalize_key


class IngestError(ValueError):
    """Raised when workbook ingestion fails."""


def _ensure_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - import guard
        raise ImportError(
            "openpyxl is required for arch pack. Install with: pip install onetool-mcp[dev]"
        ) from exc
    return openpyxl


def discover_workbooks(*, input_path: str) -> list[Path]:
    """Discover workbook files from file, directory, or glob path."""
    resolved = resolve_cwd_path(input_path)
    if resolved.is_file():
        if resolved.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise IngestError(f"input_path must point to .xlsx/.xlsm file: {resolved}")
        return [resolved]

    candidates: list[Path]
    raw = Path(input_path)
    if any(token in input_path for token in ("*", "?", "[")):
        if raw.is_absolute():
            root = Path(raw.anchor)
            pattern = str(raw.relative_to(raw.anchor))
            candidates = sorted(
                path.resolve() for path in root.glob(pattern) if path.is_file()
            )
        else:
            root = resolve_cwd_path(".")
            pattern = input_path
            candidates = sorted(path.resolve() for path in root.glob(pattern) if path.is_file())
    elif resolved.is_dir():
        candidates = sorted(path.resolve() for path in resolved.glob("*.xls*") if path.is_file())
    else:
        raise IngestError(f"No workbook(s) found for input_path: {input_path}")

    workbooks = [path for path in candidates if path.suffix.lower() in {".xlsx", ".xlsm"}]
    if not workbooks:
        raise IngestError(f"No workbook(s) found for input_path: {input_path}")
    return workbooks


def _sheet_rows_from_workbook(
    *,
    workbook: Any,
    workbook_path: Path,
    sheet_name: str,
) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    header_values = [normalize_key(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if not any(header_values):
        return []
    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        parsed: dict[str, Any] = {}
        for idx, header in enumerate(header_values):
            if not header:
                continue
            value = normalize_cell(row[idx] if idx < len(row) else None)
            if value is None:
                continue
            if isinstance(value, str) and not value:
                continue
            parsed[header] = value
        if parsed:
            parsed["_sheet_row"] = row_index
            parsed["_source_file"] = str(workbook_path)
            rows.append(parsed)
    return rows


def ingest_workbooks(*, workbook_paths: list[Path]) -> dict[str, list[dict[str, Any]]]:
    """Read entity rows from workbook list."""
    openpyxl = _ensure_openpyxl()
    merged: dict[str, list[dict[str, Any]]] = {sheet: [] for sheet in SHEETS}
    for workbook_path in workbook_paths:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        try:
            for sheet in SHEETS:
                merged[sheet].extend(
                    _sheet_rows_from_workbook(
                        workbook=workbook,
                        workbook_path=workbook_path,
                        sheet_name=sheet,
                    )
                )
        finally:
            workbook.close()
    return merged


def ingest_input(*, input_path: str) -> tuple[list[Path], dict[str, list[dict[str, Any]]]]:
    """Resolve input and return discovered workbooks and parsed entities."""
    workbooks = discover_workbooks(input_path=input_path)
    entities = ingest_workbooks(workbook_paths=workbooks)
    return workbooks, entities


__all__ = ["IngestError", "discover_workbooks", "ingest_input", "ingest_workbooks"]
