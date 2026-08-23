"""Excel adapter for architecture schema v3."""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from pydantic import ValidationError

from .model import Architecture
from .validate import validate
from .yamlio import dump_architecture, format_data_path, load_architecture

SHEETS = (
    "Architecture",
    "Milestones",
    "Timelines",
    "Systems",
    "Subsystems",
    "Components",
    "Users",
    "Interfaces",
    "Relationships",
)
COLLECTIONS = {
    "Milestones": "milestones",
    "Systems": "systems",
    "Subsystems": "subsystems",
    "Components": "components",
    "Users": "users",
    "Interfaces": "interfaces",
    "Relationships": "relationships",
}
HEADERS = {
    "Architecture": ("schema_version",),
    "Milestones": ("id", "name", "description", "tags"),
    "Timelines": ("timeline", "milestone"),
    "Systems": ("id", "name", "from", "until", "description", "tags"),
    "Subsystems": (
        "id",
        "name",
        "system",
        "from",
        "until",
        "description",
        "tags",
    ),
    "Components": (
        "id",
        "name",
        "subsystem",
        "from",
        "until",
        "description",
        "tags",
    ),
    "Users": ("id", "name", "from", "until", "description", "tags"),
    "Interfaces": (
        "id",
        "name",
        "provider",
        "consumer",
        "call_direction",
        "data_flow",
        "from",
        "until",
        "description",
        "tags",
    ),
    "Relationships": (
        "id",
        "source",
        "action",
        "target",
        "from",
        "until",
        "description",
        "tags",
    ),
}
ID_FIELDS = {
    "id",
    "timeline",
    "milestone",
    "system",
    "subsystem",
    "provider",
    "consumer",
    "source",
    "target",
    "from",
    "until",
}
ENUM_FIELDS = {"call_direction", "data_flow"}
LIST_FIELDS = {"tags"}
PROPERTY_SHEETS = set(COLLECTIONS)
MILESTONE_VALIDATION_NAME = "arch_milestones"


class WorkbookError(ValueError):
    """Raised when a workbook cannot represent a valid architecture."""


@dataclass(frozen=True)
class _CellLocation:
    path: Path
    sheet: str
    row: int
    column: str
    field: str

    def render(self, message: str) -> str:
        return (
            f"{self.path}: sheet {self.sheet!r}, row {self.row}, "
            f"column {self.column}, field {self.field!r}: {message}"
        )


@dataclass
class _SheetRows:
    headers: list[str]
    original_headers: list[str]
    rows: list[tuple[int, dict[str, Any]]]
    locations: dict[tuple[int, str], _CellLocation]


def _normalise(value: object) -> str:
    text = "" if value is None else str(value).strip().lower()
    return "_".join(re.sub(r"[^\w]+", " ", text).split())


def _identifier(value: object) -> str:
    if isinstance(value, bool):
        raise ValueError("boolean values are not valid identifiers")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        raise ValueError("decimal values are not valid identifiers")
    text = str(value).strip()
    if not text:
        raise ValueError("identifiers must not be blank")
    return text


def _list_value(value: object) -> list[str]:
    if not isinstance(value, str):
        raise ValueError("list values must be text")
    text = value.strip()
    if text.startswith("[") != text.endswith("]"):
        raise ValueError("list cells must use matching square brackets")
    if text.startswith("["):
        text = text[1:-1]
        if not text.strip():
            return []
        values = [item.strip() for item in text.split(";")]
    else:
        values = [text]
    if any(not item or any(char in item for char in ";[]\n\r") for item in values):
        raise ValueError(
            "list items must be nonblank and contain no brackets, semicolons, or newlines"
        )
    return values


def _property_value(value: object) -> str | list[str]:
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("[") or text.endswith("]"):
            return _list_value(text)
        if not text:
            raise ValueError("property values must not be blank")
        return text
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        raise ValueError("decimal property values must be whole numbers")
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise ValueError(f"unsupported property value type {type(value).__name__}")


def _cell_value(value: object, field: str, *, property_column: bool) -> Any:
    if property_column:
        return _property_value(value)
    if field in ID_FIELDS:
        return _identifier(value)
    if field in LIST_FIELDS:
        return _list_value(value)
    if field in ENUM_FIELDS and isinstance(value, str):
        return value.strip().lower()
    return value.strip() if isinstance(value, str) else value


def _read_sheet(
    path: Path, worksheet: Any, canonical: str
) -> tuple[_SheetRows, list[str]]:
    issues: list[str] = []
    header_cells = list(next(worksheet.iter_rows(min_row=1, max_row=1), ()))
    headers = [_normalise(cell.value) for cell in header_cells]
    originals = [
        "" if cell.value is None else str(cell.value).strip() for cell in header_cells
    ]
    required = set(HEADERS[canonical])
    counts = Counter(header for header in headers if header)
    for header, count in counts.items():
        if count > 1:
            columns = ", ".join(
                cell.column_letter
                for cell, item in zip(header_cells, headers, strict=True)
                if item == header
            )
            issues.append(
                f"{path}: sheet {canonical!r}, row 1, columns {columns}: duplicate header {header!r}"
            )
    for header in sorted(required - set(headers)):
        issues.append(
            f"{path}: sheet {canonical!r}, row 1, field {header!r}: required column is missing"
        )
    if canonical not in PROPERTY_SHEETS:
        for cell, header, original in zip(
            header_cells, headers, originals, strict=True
        ):
            if header and header not in required:
                issues.append(
                    f"{path}: sheet {canonical!r}, row 1, column {cell.column_letter}, "
                    f"field {original!r}: unknown column"
                )

    rows: list[tuple[int, dict[str, Any]]] = []
    locations: dict[tuple[int, str], _CellLocation] = {}
    for row_number, cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
        populated = [
            cell
            for cell in cells
            if cell.value is not None
            and not (isinstance(cell.value, str) and not cell.value.strip())
        ]
        if not populated:
            continue
        row: dict[str, Any] = {}
        for index, cell in enumerate(cells):
            if cell.value is None or (
                isinstance(cell.value, str) and not cell.value.strip()
            ):
                continue
            header = headers[index] if index < len(headers) else ""
            original = originals[index] if index < len(originals) else ""
            field = original or header or "<blank>"
            location = _CellLocation(
                path, canonical, row_number, cell.column_letter, field
            )
            if not header:
                issues.append(location.render("populated cell has no header"))
                continue
            if cell.data_type == "f":
                issues.append(location.render("formulas are not allowed"))
                continue
            locations[(len(rows), header)] = location
            try:
                row[header if header in required else original] = _cell_value(
                    cell.value,
                    header,
                    property_column=header not in required,
                )
            except ValueError as exc:
                issues.append(location.render(str(exc)))
        rows.append((row_number, row))
    return _SheetRows(headers, originals, rows, locations), issues


def _source_location(
    path: Path,
    source_map: dict[str, _CellLocation],
    data_path: tuple[str | int, ...],
) -> _CellLocation:
    for length in range(len(data_path), 0, -1):
        candidate = format_data_path(data_path[:length])
        if candidate in source_map:
            return source_map[candidate]
    field = str(data_path[-1]) if data_path else "<root>"
    return _CellLocation(path, "<workbook>", 1, "?", field)


def _architecture_raw(
    path: Path, sheets: dict[str, _SheetRows]
) -> tuple[dict[str, Any], dict[str, _CellLocation], list[str]]:
    issues: list[str] = []
    sources: dict[str, _CellLocation] = {}
    architecture_rows = sheets["Architecture"].rows
    if len(architecture_rows) != 1:
        issues.append(
            f"{path}: sheet 'Architecture': expected exactly one populated data row"
        )
    schema_version = (
        architecture_rows[0][1].get("schema_version") if architecture_rows else None
    )
    if architecture_rows:
        location = sheets["Architecture"].locations.get((0, "schema_version"))
        if location:
            sources["schema_version"] = location

    raw: dict[str, Any] = {"schema_version": schema_version, "milestones": []}
    for sheet, collection in COLLECTIONS.items():
        collection_rows: list[dict[str, Any]] = []
        reserved = set(HEADERS[sheet])
        for index, (_row_number, row) in enumerate(sheets[sheet].rows):
            item = {
                key: value for key, value in row.items() if _normalise(key) in reserved
            }
            properties = {
                key: value
                for key, value in row.items()
                if _normalise(key) not in reserved
            }
            if properties:
                item["properties"] = properties
            collection_rows.append(item)
            for (row_index, header), location in sheets[sheet].locations.items():
                if row_index != index:
                    continue
                key = header if header in reserved else "properties"
                sources[f"{collection}[{index}].{key}"] = location
            if "id" in item:
                sources.setdefault(
                    f"{collection}[{index}]",
                    sources.get(
                        f"{collection}[{index}].id",
                        _CellLocation(path, sheet, _row_number, "?", "id"),
                    ),
                )
        raw[collection] = collection_rows

    timeline_rows = sheets["Timelines"].rows
    timelines: list[dict[str, Any]] = []
    seen: set[str] = set()
    current: str | None = None
    for row_index, (row_number, row) in enumerate(timeline_rows):
        timeline = row.get("timeline")
        milestone = row.get("milestone")
        if timeline is None or milestone is None:
            for field, value in (("timeline", timeline), ("milestone", milestone)):
                if value is None:
                    location = sheets["Timelines"].locations.get((row_index, field))
                    message = f"required field {field!r} is missing"
                    issues.append(
                        location.render(message)
                        if location
                        else f"{path}: sheet 'Timelines', row {row_number}, field {field!r}: {message}"
                    )
            continue
        if timeline != current:
            if timeline in seen:
                location = sheets["Timelines"].locations.get((row_index, "timeline"))
                message = (
                    f"timeline {timeline!r} appears in more than one contiguous block"
                )
                issues.append(
                    location.render(message)
                    if location
                    else f"{path}: sheet 'Timelines', row {row_number}: {message}"
                )
            seen.add(timeline)
            timelines.append({"id": timeline, "milestones": []})
            current = timeline
        timeline_index = len(timelines) - 1
        milestone_index = len(timelines[timeline_index]["milestones"])
        timelines[timeline_index]["milestones"].append(milestone)
        for header, model_path in (
            ("timeline", f"timelines[{timeline_index}].id"),
            ("milestone", f"timelines[{timeline_index}].milestones[{milestone_index}]"),
        ):
            location = sheets["Timelines"].locations.get((row_index, header))
            if location:
                sources[model_path] = location
    raw["timelines"] = timelines or None
    return raw, sources, issues


def _validation_issues(
    path: Path,
    architecture: Architecture,
    source_map: dict[str, _CellLocation],
) -> list[str]:
    issues: list[str] = []
    duplicate_paths: set[str] = set()
    for finding in validate(architecture):
        if finding.severity != "error":
            continue
        location = source_map.get(finding.path)
        if finding.code == "duplicate_id":
            match = re.match(r"(\w+)\[(\d+)]", finding.path)
            if match:
                collection, index_text = match.groups()
                duplicate_id = getattr(architecture, collection)[int(index_text)].id
                for index, row in enumerate(getattr(architecture, collection)):
                    if row.id == duplicate_id:
                        duplicate_paths.add(f"{collection}[{index}].id")
                continue
        if location is None:
            location = _source_location(path, source_map, ())
        issues.append(location.render(f"{finding.code}: {finding.message}"))
    for data_path in sorted(duplicate_paths):
        location = source_map.get(data_path)
        if location:
            issues.append(
                location.render("duplicate_id: identity has conflicting declarations")
            )
    return issues


def read_workbook(path: Path) -> Architecture:
    """Read one `.xlsx` workbook into a validated architecture model."""
    if path.suffix.lower() != ".xlsx":
        raise WorkbookError(f"Excel architecture input must use .xlsx: {path}")
    try:
        workbook = load_workbook(path, data_only=False)
    except (OSError, ValueError) as exc:
        raise WorkbookError(f"Unable to read workbook {path}: {exc}") from exc
    try:
        by_name = {name.casefold(): name for name in workbook.sheetnames}
        missing = [name for name in SHEETS if name.casefold() not in by_name]
        if missing:
            raise WorkbookError(
                f"{path}: missing required sheets: {', '.join(missing)}"
            )
        parsed: dict[str, _SheetRows] = {}
        issues: list[str] = []
        for canonical in SHEETS:
            sheet_rows, sheet_issues = _read_sheet(
                path, workbook[by_name[canonical.casefold()]], canonical
            )
            parsed[canonical] = sheet_rows
            issues.extend(sheet_issues)
        raw, sources, structure_issues = _architecture_raw(path, parsed)
        issues.extend(structure_issues)
        try:
            architecture = Architecture.model_validate(raw)
        except ValidationError as exc:
            for detail in exc.errors(include_url=False):
                data_path = tuple(
                    part for part in detail["loc"] if isinstance(part, (str, int))
                )
                location = _source_location(path, sources, data_path)
                issues.append(
                    location.render(f"{format_data_path(data_path)}: {detail['msg']}")
                )
            architecture = None
        if architecture is not None:
            issues.extend(_validation_issues(path, architecture, sources))
        if issues:
            raise WorkbookError("\n".join(dict.fromkeys(issues)))
        assert architecture is not None
        return architecture
    finally:
        workbook.close()


def _rows(architecture: Architecture) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = {
        "Architecture": [{"schema_version": 3}],
        "Timelines": [
            {"timeline": timeline.id, "milestone": milestone}
            for timeline in architecture.timelines or []
            for milestone in timeline.milestones
        ],
    }
    for sheet, collection in COLLECTIONS.items():
        values: list[dict[str, Any]] = []
        for item in getattr(architecture, collection):
            row = item.model_dump(mode="python", by_alias=True, exclude_none=True)
            properties = row.pop("properties", {})
            collisions = {
                key for key in properties if _normalise(key) in set(HEADERS[sheet])
            }
            if collisions:
                names = ", ".join(sorted(collisions, key=str.casefold))
                raise WorkbookError(
                    f"sheet {sheet!r} properties collide with reserved columns: {names}"
                )
            row.update(properties)
            values.append(row)
        result[sheet] = values
    return result


def _property_headers(sheet: str, rows: list[dict[str, Any]]) -> list[str]:
    reserved = set(HEADERS[sheet])
    properties: dict[str, str] = {}
    for row in rows:
        for key in row:
            normalised = _normalise(key)
            if normalised in reserved:
                continue
            previous = properties.setdefault(normalised, key)
            if previous != key:
                raise WorkbookError(
                    f"sheet {sheet!r} has colliding property names {previous!r} and {key!r}"
                )
    return sorted(
        properties.values(), key=lambda item: (_normalise(item), item.casefold())
    )


def _excel_value(value: Any) -> Any:
    if isinstance(value, list):
        return "[" + ";".join(value) + "]"
    return value


def _table_name(sheet: str) -> str:
    return "Arch" + re.sub(r"[^A-Za-z0-9]", "", sheet)


def _add_validations(
    worksheet: Any,
    headers: list[str],
    end_row: int,
    only: set[str] | None = None,
) -> None:
    by_header = {_normalise(header): index + 1 for index, header in enumerate(headers)}
    for field in ("milestone", "from", "until"):
        if field not in by_header or (only is not None and field not in only):
            continue
        validation = DataValidation(
            type="list", formula1=f"={MILESTONE_VALIDATION_NAME}", allow_blank=True
        )
        worksheet.add_data_validation(validation)
        column = worksheet.cell(1, by_header[field]).column_letter
        validation.add(f"{column}2:{column}{end_row}")
    choices = '"provider_to_consumer,consumer_to_provider,bidirectional,unspecified"'
    for field in ENUM_FIELDS:
        if field not in by_header or (only is not None and field not in only):
            continue
        validation = DataValidation(type="list", formula1=choices, allow_blank=False)
        worksheet.add_data_validation(validation)
        column = worksheet.cell(1, by_header[field]).column_letter
        validation.add(f"{column}2:{column}{end_row}")


def _new_workbook(architecture: Architecture) -> Workbook:
    workbook = Workbook()
    if workbook.active is not None:
        workbook.remove(workbook.active)
    workbook.defined_names.add(
        DefinedName(
            MILESTONE_VALIDATION_NAME,
            attr_text="'Milestones'!$A$2:$A$1048576",
        )
    )
    rows_by_sheet = _rows(architecture)
    revision_fill = PatternFill("solid", fgColor="EAF2F8")
    for sheet in SHEETS:
        worksheet = workbook.create_sheet(sheet)
        rows = rows_by_sheet[sheet]
        headers = list(HEADERS[sheet]) + _property_headers(sheet, rows)
        worksheet.append(headers)
        for row in rows:
            normalised = {_normalise(key): value for key, value in row.items()}
            worksheet.append(
                [_excel_value(normalised.get(_normalise(header))) for header in headers]
            )
        if not rows:
            worksheet.append([None] * len(headers))
        end_row = max(2, len(rows) + 1)
        table = Table(
            displayName=_table_name(sheet),
            ref=f"A1:{worksheet.cell(end_row, len(headers)).coordinate}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        worksheet.freeze_panes = "A2"
        _add_validations(worksheet, headers, end_row)
        if sheet in COLLECTIONS and rows:
            ids = Counter(str(row.get("id")) for row in rows)
            for index, row in enumerate(rows, start=2):
                if ids[str(row.get("id"))] > 1:
                    for cell in worksheet[index]:
                        cell.fill = revision_fill
            id_column = headers.index("id") + 1
            letter = worksheet.cell(1, id_column).column_letter
            worksheet.conditional_formatting.add(
                f"A2:{worksheet.cell(end_row, len(headers)).coordinate}",
                FormulaRule(
                    formula=[f"COUNTIF(${letter}:${letter},${letter}2)>1"],
                    fill=revision_fill,
                ),
            )
    return workbook


def _canonical_sheet(workbook: Workbook, canonical: str) -> Any:
    matches = [
        name for name in workbook.sheetnames if name.casefold() == canonical.casefold()
    ]
    if len(matches) != 1:
        raise WorkbookError(f"existing workbook must contain one {canonical!r} sheet")
    return workbook[matches[0]]


def _update_workbook(workbook: Workbook, architecture: Architecture) -> None:
    if any(worksheet._charts or worksheet._images for worksheet in workbook.worksheets):
        raise WorkbookError(
            "existing workbook contains charts or images that openpyxl cannot preserve"
        )
    rows_by_sheet = _rows(architecture)
    workbook.defined_names.add(
        DefinedName(MILESTONE_VALIDATION_NAME, attr_text="'Milestones'!$A$2:$A$1048576")
    )
    for sheet in SHEETS:
        worksheet = _canonical_sheet(workbook, sheet)
        tables = list(worksheet.tables.values())
        if len(tables) != 1:
            raise WorkbookError(
                f"sheet {sheet!r} must contain exactly one structured table"
            )
        table = tables[0]
        min_col, min_row, max_col, old_end_row = range_boundaries(table.ref)
        if min_row != 1:
            raise WorkbookError(f"sheet {sheet!r} table must start on row 1")
        headers = [
            str(worksheet.cell(1, column).value or "").strip()
            for column in range(min_col, max_col + 1)
        ]
        normalised_headers = [_normalise(header) for header in headers]
        missing = set(HEADERS[sheet]) - set(normalised_headers)
        if missing:
            raise WorkbookError(
                f"sheet {sheet!r} is missing reserved columns: {', '.join(sorted(missing))}"
            )
        rows = rows_by_sheet[sheet]
        required_properties = {
            _normalise(item) for item in _property_headers(sheet, rows)
        }
        available_properties = set(normalised_headers) - set(HEADERS[sheet])
        if missing_properties := required_properties - available_properties:
            raise WorkbookError(
                f"sheet {sheet!r} has no columns for properties: {', '.join(sorted(missing_properties))}"
            )
        end_row = max(2, len(rows) + 1)
        for row_number in range(2, max(old_end_row, end_row) + 1):
            for column in range(min_col, max_col + 1):
                worksheet.cell(row_number, column).value = None
        for row_number, row in enumerate(rows, start=2):
            normalised = {_normalise(key): value for key, value in row.items()}
            for offset, header in enumerate(normalised_headers):
                worksheet.cell(row_number, min_col + offset).value = _excel_value(
                    normalised.get(header)
                )
        table.ref = f"{worksheet.cell(1, min_col).coordinate}:{worksheet.cell(end_row, max_col).coordinate}"
        controlled = {
            field: normalised_headers.index(field) + min_col
            for field in ("milestone", "from", "until", *ENUM_FIELDS)
            if field in normalised_headers
        }
        matched: set[str] = set()
        for validation in worksheet.data_validations.dataValidation:
            for cell_range in list(validation.ranges.ranges):
                for field, column in controlled.items():
                    if (
                        cell_range.min_col <= column <= cell_range.max_col
                        and cell_range.min_row >= 2
                    ):
                        letter = worksheet.cell(1, column).column_letter
                        validation.sqref = f"{letter}2:{letter}{end_row}"
                        matched.add(field)
        if missing_validations := set(controlled) - matched:
            _add_validations(worksheet, headers, end_row, missing_validations)


def write_workbook(architecture: Architecture, path: Path) -> None:
    """Write a new workbook or atomically update an existing workbook in place."""
    if path.suffix.lower() != ".xlsx":
        raise WorkbookError(f"Excel architecture output must use .xlsx: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(path) if path.exists() else _new_workbook(architecture)
    try:
        if path.exists():
            _update_workbook(workbook, architecture)
        with NamedTemporaryFile(
            dir=path.parent, delete=False, suffix=".xlsx"
        ) as handle:
            temporary = Path(handle.name)
        try:
            workbook.save(temporary)
            temporary.replace(path)
        finally:
            temporary.unlink(missing_ok=True)
    finally:
        workbook.close()


def generate_template(path: Path) -> None:
    """Generate an empty canonical workbook without overwriting an existing file."""
    if path.exists():
        raise FileExistsError(f"workbook already exists: {path}")
    architecture = Architecture(
        schema_version=3,
        milestones=[],
        systems=[],
        subsystems=[],
        components=[],
        users=[],
        interfaces=[],
        relationships=[],
    )
    write_workbook(architecture, path)


def import_workbook(workbook_path: Path, yaml_path: Path) -> dict[str, Any]:
    """Validate a workbook and atomically replace its canonical YAML target."""
    architecture = read_workbook(workbook_path)
    dump_architecture(architecture, yaml_path)
    return {"ok": True, "input_path": str(workbook_path), "path": str(yaml_path)}


def export_workbook(yaml_path: Path, workbook_path: Path) -> dict[str, Any]:
    """Load canonical YAML and write its Excel representation."""
    architecture = load_architecture(yaml_path)
    errors = [
        finding for finding in validate(architecture) if finding.severity == "error"
    ]
    if errors:
        codes = ", ".join(dict.fromkeys(finding.code for finding in errors))
        raise WorkbookError(f"architecture has validation errors: {codes}")
    write_workbook(architecture, workbook_path)
    return {"ok": True, "input_path": str(yaml_path), "path": str(workbook_path)}
