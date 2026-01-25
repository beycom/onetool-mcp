"""Excel workbook to Markdown converter.

Converts XLSX spreadsheets to Markdown with:
- Streaming row processing via openpyxl read_only mode
- Sheet-based sections
- Optional formula extraction
- YAML frontmatter and TOC generation
"""

from __future__ import annotations

from pathlib import Path  # noqa: TC003 (used at runtime)
from typing import Any

try:
    from openpyxl import load_workbook  # type: ignore[import-untyped]
except ImportError as e:
    raise ImportError(
        "openpyxl is required for convert. Install with: pip install openpyxl"
    ) from e

from ot_tools._convert.utils import (
    IncrementalWriter,
    compute_file_checksum,
    get_mtime_iso,
    normalise_whitespace,
    write_toc_file,
)


def convert_excel(
    input_path: Path,
    output_dir: Path,
    source_rel: str,
    *,
    include_formulas: bool = False,
) -> dict[str, Any]:
    """Convert Excel workbook to Markdown.

    Args:
        input_path: Path to XLSX file
        output_dir: Directory for output files
        source_rel: Relative path to source for frontmatter
        include_formulas: Include cell formulas as comments

    Returns:
        Dict with 'output', 'sheets', 'rows' keys
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load workbook once in read-only mode for streaming
    # - data_only=True: get computed values (no formulas)
    # - data_only=False: get formulas as cell values (when include_formulas=True)
    wb = load_workbook(input_path, read_only=True, data_only=not include_formulas)

    # Get metadata for frontmatter
    checksum = compute_file_checksum(input_path)
    mtime = get_mtime_iso(input_path)
    total_sheets = len(wb.sheetnames)

    writer = IncrementalWriter()
    total_rows = 0

    # Process each sheet (single workbook - no double loading)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = _process_sheet(writer, sheet_name, ws, include_formulas)
        total_rows += rows

    wb.close()

    # Write main output (pure content, no frontmatter - line numbers start at 1)
    content = normalise_whitespace(writer.get_content())
    output_path = output_dir / f"{input_path.stem}.md"
    output_path.write_text(content, encoding="utf-8")

    # Write separate TOC file (includes frontmatter)
    headings = writer.get_headings()
    toc_path = write_toc_file(
        headings=headings,
        output_dir=output_dir,
        stem=input_path.stem,
        source=source_rel,
        converted=mtime,
        pages=total_sheets,
        checksum=checksum,
    )

    return {
        "output": str(output_path),
        "toc": str(toc_path),
        "sheets": total_sheets,
        "rows": total_rows,
    }


def _process_sheet(
    writer: IncrementalWriter,
    sheet_name: str,
    ws: Any,
    include_formulas: bool,
) -> int:
    """Process a single worksheet with streaming (O(1) memory for row data).

    When include_formulas=True, the workbook was loaded with data_only=False,
    so formula cells contain the formula string as their value.

    Returns:
        Number of rows processed
    """
    writer.write_heading(2, f"Sheet: {sheet_name}")

    # First pass: count max columns (streaming, no data storage)
    max_cols = 0
    row_count = 0
    for row in ws.iter_rows():
        max_cols = max(max_cols, len(row))
        row_count += 1

    if row_count == 0:
        writer.write("(empty sheet)\n\n")
        return 0

    # Second pass: stream rows directly to writer
    rows_iter = iter(ws.iter_rows())

    # Get header (first row)
    first_row = next(rows_iter)
    header = [str(cell.value) if cell.value is not None else "" for cell in first_row]
    # Pad header to max_cols
    while len(header) < max_cols:
        header.append("")

    # Write header
    writer.write("| " + " | ".join(_escape_pipe(c) for c in header) + " |\n")
    writer.write("| " + " | ".join("---" for _ in header) + " |\n")

    # Collect formulas as we go (just formula tuples, not full row data)
    # Format: (col_letter, row_num, formula_string)
    formulas: list[tuple[str, int, str]] = []

    # Check first row for formulas (cell values are formulas when include_formulas=True)
    if include_formulas:
        for j, cell in enumerate(first_row):
            try:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas.append((_col_letter(j + 1), 1, value))
            except Exception:
                pass

    # Stream remaining rows directly to writer
    current_row = 2  # 1-indexed, header was row 1
    for row in rows_iter:
        row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
        # Pad row to max_cols
        while len(row_values) < max_cols:
            row_values.append("")

        writer.write("| " + " | ".join(_escape_pipe(c) for c in row_values[:len(header)]) + " |\n")

        # Track formulas for this row (cell values are formulas when include_formulas=True)
        if include_formulas:
            for j, cell in enumerate(row):
                try:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas.append((_col_letter(j + 1), current_row, value))
                except Exception:
                    pass

        current_row += 1

    writer.write("\n")

    # Add formulas section if any formulas found
    if formulas:
        writer.write("**Formulas:**\n\n")
        writer.write("```\n")
        for col_letter, row_num, formula in formulas:
            writer.write(f"{col_letter}{row_num}: {formula}\n")
        writer.write("```\n\n")

    return row_count


def _escape_pipe(text: str) -> str:
    """Escape pipe characters for Markdown tables."""
    return text.replace("|", "\\|").replace("\n", " ")


def _col_letter(n: int) -> str:
    """Convert column number to letter (1=A, 2=B, ..., 27=AA)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result
